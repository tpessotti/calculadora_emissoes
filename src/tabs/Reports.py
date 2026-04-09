import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
import math
import json
import sys
import os
from datetime import datetime
from io import BytesIO

# Ensure core is importable
_root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _root_dir not in sys.path:
    sys.path.insert(0, _root_dir)

from core.context import AppContext
from core.units import (
    normalize_unit as _norm_unit, co2e_label, co2e_intensity_label,
    convert_co2e, get_default_mass_unit_from_session,
)
from calculations import EmissionCalculator
from tabs.report_exports import (
    render_report_config, get_report_config, init_report_config,
    generate_md_painel_geral, generate_md_inventario, generate_md_ifrs,
    generate_md_analise_unidade, generate_md_comparativo,
    generate_pdf_painel_geral, generate_pdf_analise_unidade, generate_pdf_comparativo,
    render_download_bar,
    _pdf_safe, _fmt_num,
)

REPORTLAB_AVAILABLE = False  # PDF generation removed; markdown-only exports

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─── Paleta de cores ────────────────────────────────────────────────
_COLORS = {
    "primary":   "#0F766E", "secondary": "#0284C7", "accent":    "#7C3AED",
    "success":   "#059669", "warning":   "#D97706", "danger":    "#DC2626",
    "info":      "#0891B2", "bg_card":   "#F8FAFC", "border":    "#E2E8F0",
    "text":      "#1E293B", "muted":     "#94A3B8",
    "scope1":    "#EF4444", "scope2":    "#F59E0B", "scope3":    "#3B82F6",
}

# ─── Defaults do questionário ───────────────────────────────────────
_Q_DEFAULTS = {
    # Identificação
    "q_empresa_nome": "",
    "q_empresa_cnpj": "",
    "q_empresa_setor": "",
    "q_empresa_pais": "Brasil",
    "q_empresa_responsavel": "",
    "q_empresa_cargo_responsavel": "",
    "q_empresa_email": "",
    "q_periodo_inicio": "",
    "q_periodo_fim": "",
    "q_moeda": "BRL (R$)",
    "q_consolidacao": "Controle operacional",
    # Governança
    "q_gov_orgao": "",
    "q_gov_frequencia": "Trimestral",
    "q_gov_comite": False,
    "q_gov_comite_nome": "",
    "q_gov_competencias": "",
    "q_gov_integracao_estrategia": "",
    "q_gov_remuneracao_vinculada": False,
    "q_gov_remuneracao_detalhes": "",
    # Estratégia
    "q_est_riscos_fisicos": [],
    "q_est_riscos_transicao": [],
    "q_est_horizonte_curto": "0-2 anos",
    "q_est_horizonte_medio": "2-5 anos",
    "q_est_horizonte_longo": "5-30 anos",
    "q_est_cenarios": False,
    "q_est_cenarios_detalhes": "",
    "q_est_impacto_receita": "",
    "q_est_impacto_custos": "",
    "q_est_impacto_ativos": "",
    "q_est_oportunidades_desc": "",
    # Gestão de Riscos
    "q_risk_processo": "",
    "q_risk_frequencia": "Anual",
    "q_risk_integrado": False,
    "q_risk_integrado_desc": "",
    "q_risk_mitigacao": "",
    # Metas
    "q_meta_possui": False,
    "q_meta_tipo": "Absoluta (tCO₂e)",
    "q_meta_base_ano": 2024,
    "q_meta_alvo_ano": 2030,
    "q_meta_reducao_pct": 30,
    "q_meta_sbti": False,
    "q_meta_net_zero_ano": "",
    "q_meta_intermediarias": "",
    # Plano de Transição
    "q_trans_possui_plano": False,
    "q_trans_acoes": "",
    "q_trans_investimento": "",
    "q_trans_tecnologias": "",
    "q_trans_dependencias": "",
    # Verificação
    "q_ver_assegurado": False,
    "q_ver_tipo": "Limitada",
    "q_ver_auditor": "",
    "q_ver_norma": "ISO 14064-3",
    # Informações Adicionais
    "q_add_offsets": False,
    "q_add_offsets_desc": "",
    "q_add_preco_interno": False,
    "q_add_preco_valor": 0,
    "q_add_notas": "",
}


class ReportsTab:
    """Página de Análises & Reportes – IFRS S1/S2, Sankey, KPIs."""

    # ────────────────────────────────────────────────────────────────
    #  Helpers para questionário
    # ────────────────────────────────────────────────────────────────
    def _init_questionario(self):
        if "ifrs_questionario" not in st.session_state:
            st.session_state.ifrs_questionario = dict(_Q_DEFAULTS)

    def _q(self, key):
        """Retorna valor do questionário ou default."""
        self._init_questionario()
        return st.session_state.ifrs_questionario.get(key, _Q_DEFAULTS.get(key, ""))

    def _q_set(self, key, value):
        self._init_questionario()
        st.session_state.ifrs_questionario[key] = value

    def _q_preenchido(self):
        """Verifica se o questionário tem dados mínimos preenchidos."""
        return bool(self._q("q_empresa_nome"))

    # ════════════════════════════════════════════════════════════════
    #  RENDER PRINCIPAL
    # ════════════════════════════════════════════════════════════════
    def _render(self):
        self._init_questionario()
        init_report_config()

        # ── Header ──
        col_title, col_badge = st.columns([4, 1])
        with col_title:
            st.markdown(
                "<h1 style='margin-bottom:0'>📊 Análises & Reportes</h1>"
                "<p style='color:#64748B;margin-top:0'>Visualização de emissões e reportes alinhados "
                "aos padrões internacionais IFRS S1/S2</p>",
                unsafe_allow_html=True,
            )
        with col_badge:
            cfg = get_report_config()
            if cfg.get("rpt_confidencial"):
                st.markdown("<div style='background:#DC2626;color:white;padding:4px 12px;"
                            "border-radius:6px;text-align:center;margin-top:18px;"
                            "font-weight:600;font-size:0.8rem'>🔒 CONFIDENCIAL</div>",
                            unsafe_allow_html=True)
            tema = cfg.get("rpt_tema", "Padrão")
            if tema != "Padrão":
                st.caption(f"Tema: {tema}")

        # ── Customization panel ──
        render_report_config()

        if not st.session_state.unidades:
            st.info("Adicione unidades na aba **Unidades & Fluxos** para visualizar análises.", icon="ℹ️")
            self._render_ifrs_info_only()
            return

        from calculations import EmissionCalculator  # keep for reload safety
        EmissionCalculator.propagar_pegada(st.session_state.unidades, st.session_state.edges)

        tabs = st.tabs([
            "📈 Painel Geral",
            "🔀 Diagrama Sankey",
            "📑 Inventário GEE",
            "📝 Questionário IFRS",
            "📋 Reporte IFRS S1/S2",
            "📑 Análise por Unidade",
            "🔄 Comparativo Multi-Ano",
        ])

        with tabs[0]:
            self._render_painel_geral()
        with tabs[1]:
            self._render_sankey_diagram()
        with tabs[2]:
            self._render_inventario_ghg()
        with tabs[3]:
            self._render_questionario()
        with tabs[4]:
            self._render_ifrs_report()
        with tabs[5]:
            self._render_analise_por_unidade()
        with tabs[6]:
            self._render_comparativo()

    # ════════════════════════════════════════════════════════════════
    #  QUESTIONÁRIO IFRS
    # ════════════════════════════════════════════════════════════════
    def _render_questionario(self):
        st.markdown("### 📝 Questionário de Informações Corporativas – IFRS S1/S2")
        st.markdown(
            "*Preencha as informações abaixo para gerar um relatório IFRS mais completo e personalizado. "
            "Os dados são salvos automaticamente na sessão.*"
        )

        if self._q_preenchido():
            st.success(f"✅ Questionário preenchido para: **{self._q('q_empresa_nome')}**", icon="✅")
        else:
            st.warning("Preencha ao menos o nome da empresa para habilitar o relatório completo.", icon="⚠️")

        # ── 1. Identificação da Empresa ──
        with st.expander("🏢 1. Identificação da Entidade", expanded=not self._q_preenchido()):
            c1, c2 = st.columns(2)
            with c1:
                v = st.text_input("Nome da Empresa / Entidade *", value=self._q("q_empresa_nome"), key="_q_nome")
                self._q_set("q_empresa_nome", v)
                v = st.text_input("CNPJ / Registro", value=self._q("q_empresa_cnpj"), key="_q_cnpj")
                self._q_set("q_empresa_cnpj", v)
                v = st.text_input("Setor de Atuação (NACE/CNAE)", value=self._q("q_empresa_setor"), key="_q_setor")
                self._q_set("q_empresa_setor", v)
                v = st.text_input("País-sede", value=self._q("q_empresa_pais"), key="_q_pais")
                self._q_set("q_empresa_pais", v)
            with c2:
                v = st.text_input("Responsável pelo Reporte", value=self._q("q_empresa_responsavel"), key="_q_resp")
                self._q_set("q_empresa_responsavel", v)
                v = st.text_input("Cargo", value=self._q("q_empresa_cargo_responsavel"), key="_q_cargo")
                self._q_set("q_empresa_cargo_responsavel", v)
                v = st.text_input("E-mail de contato", value=self._q("q_empresa_email"), key="_q_email")
                self._q_set("q_empresa_email", v)
                v = st.selectbox("Moeda de reporte", ["BRL (R$)", "USD ($)", "EUR (€)", "GBP (£)"],
                                 index=["BRL (R$)", "USD ($)", "EUR (€)", "GBP (£)"].index(self._q("q_moeda")),
                                 key="_q_moeda")
                self._q_set("q_moeda", v)

            c1, c2 = st.columns(2)
            with c1:
                v = st.text_input("Início do período de reporte (ex: 01/01/2025)", value=self._q("q_periodo_inicio"), key="_q_per_ini")
                self._q_set("q_periodo_inicio", v)
            with c2:
                v = st.text_input("Fim do período de reporte (ex: 31/12/2025)", value=self._q("q_periodo_fim"), key="_q_per_fim")
                self._q_set("q_periodo_fim", v)

            v = st.selectbox("Abordagem de consolidação (IFRS S1 §B25-B28)",
                             ["Controle operacional", "Controle financeiro", "Participação societária"],
                             index=["Controle operacional", "Controle financeiro", "Participação societária"].index(self._q("q_consolidacao")),
                             key="_q_consol")
            self._q_set("q_consolidacao", v)

        # ── 2. Governança ──
        with st.expander("🏛️ 2. Governança Climática (IFRS S2 §5-12)"):
            v = st.text_area(
                "Qual órgão/pessoa é responsável pela supervisão de riscos climáticos? (§6a)",
                value=self._q("q_gov_orgao"), key="_q_gov_orgao", height=80,
                placeholder="Ex: Conselho de Administração, com suporte do Comitê de Sustentabilidade"
            )
            self._q_set("q_gov_orgao", v)

            v = st.selectbox("Frequência de informação ao órgão de governança (§6b)",
                             ["Mensal", "Trimestral", "Semestral", "Anual", "Ad hoc"],
                             index=["Mensal", "Trimestral", "Semestral", "Anual", "Ad hoc"].index(self._q("q_gov_frequencia")),
                             key="_q_gov_freq")
            self._q_set("q_gov_frequencia", v)

            c1, c2 = st.columns(2)
            with c1:
                v = st.checkbox("Existe comitê dedicado à sustentabilidade/clima? (§6c)",
                                value=self._q("q_gov_comite"), key="_q_gov_comite")
                self._q_set("q_gov_comite", v)
            with c2:
                if self._q("q_gov_comite"):
                    v = st.text_input("Nome do comitê", value=self._q("q_gov_comite_nome"), key="_q_gov_comite_nm")
                    self._q_set("q_gov_comite_nome", v)

            v = st.text_area(
                "Competências e expertise climática do órgão de governança (§6d)",
                value=self._q("q_gov_competencias"), key="_q_gov_comp", height=80,
                placeholder="Descreva as qualificações e experiência em sustentabilidade..."
            )
            self._q_set("q_gov_competencias", v)

            v = st.text_area(
                "Como os riscos climáticos são integrados na estratégia de negócios? (§8)",
                value=self._q("q_gov_integracao_estrategia"), key="_q_gov_integr", height=80,
                placeholder="Ex: Riscos climáticos são avaliados em cada decisão de investimento acima de R$1M..."
            )
            self._q_set("q_gov_integracao_estrategia", v)

            c1, c2 = st.columns(2)
            with c1:
                v = st.checkbox("Remuneração vinculada a metas climáticas? (§10)",
                                value=self._q("q_gov_remuneracao_vinculada"), key="_q_gov_rem")
                self._q_set("q_gov_remuneracao_vinculada", v)
            with c2:
                if self._q("q_gov_remuneracao_vinculada"):
                    v = st.text_input("Detalhes da vinculação", value=self._q("q_gov_remuneracao_detalhes"), key="_q_gov_rem_det")
                    self._q_set("q_gov_remuneracao_detalhes", v)

        # ── 3. Estratégia ──
        with st.expander("📐 3. Estratégia Climática (IFRS S2 §13-22)"):
            st.markdown("**Riscos Físicos identificados pela organização (§13a):**")
            riscos_fisicos_opts = [
                "Eventos extremos (tempestades, inundações, secas)",
                "Aumento do nível do mar",
                "Aumento de temperatura média",
                "Escassez hídrica",
                "Incêndios florestais",
                "Mudanças em padrões de precipitação",
            ]
            v = st.multiselect("Selecione os riscos físicos aplicáveis:", riscos_fisicos_opts,
                               default=self._q("q_est_riscos_fisicos"), key="_q_est_rf")
            self._q_set("q_est_riscos_fisicos", v)

            st.markdown("**Riscos de Transição identificados (§13b):**")
            riscos_transicao_opts = [
                "Regulação de carbono / CBAM",
                "Mudanças na demanda de mercado",
                "Obsolescência tecnológica",
                "Risco reputacional",
                "Litígios climáticos",
                "Aumento de custos de energia",
                "Requisitos de divulgação (ESG)",
            ]
            v = st.multiselect("Selecione os riscos de transição aplicáveis:", riscos_transicao_opts,
                               default=self._q("q_est_riscos_transicao"), key="_q_est_rt")
            self._q_set("q_est_riscos_transicao", v)

            st.markdown("**Definição de horizontes temporais (§15):**")
            c1, c2, c3 = st.columns(3)
            with c1:
                v = st.text_input("Curto prazo", value=self._q("q_est_horizonte_curto"), key="_q_est_hc")
                self._q_set("q_est_horizonte_curto", v)
            with c2:
                v = st.text_input("Médio prazo", value=self._q("q_est_horizonte_medio"), key="_q_est_hm")
                self._q_set("q_est_horizonte_medio", v)
            with c3:
                v = st.text_input("Longo prazo", value=self._q("q_est_horizonte_longo"), key="_q_est_hl")
                self._q_set("q_est_horizonte_longo", v)

            c1, c2 = st.columns(2)
            with c1:
                v = st.checkbox("Utiliza análise de cenários climáticos? (§22)", value=self._q("q_est_cenarios"), key="_q_est_cen")
                self._q_set("q_est_cenarios", v)
            with c2:
                if self._q("q_est_cenarios"):
                    v = st.text_input("Cenários utilizados (ex: IEA NZE, SSP2-4.5)", value=self._q("q_est_cenarios_detalhes"), key="_q_est_cen_det")
                    self._q_set("q_est_cenarios_detalhes", v)

            st.markdown("**Impacto financeiro estimado (§21):**")
            c1, c2, c3 = st.columns(3)
            with c1:
                v = st.text_input("Impacto em receitas", value=self._q("q_est_impacto_receita"), key="_q_est_rec",
                                  placeholder="Ex: Risco de -5% até 2030")
                self._q_set("q_est_impacto_receita", v)
            with c2:
                v = st.text_input("Impacto em custos operacionais", value=self._q("q_est_impacto_custos"), key="_q_est_cust",
                                  placeholder="Ex: +€2M/ano em taxação")
                self._q_set("q_est_impacto_custos", v)
            with c3:
                v = st.text_input("Impacto em ativos/capital", value=self._q("q_est_impacto_ativos"), key="_q_est_ativ",
                                  placeholder="Ex: Reavaliação de -3%")
                self._q_set("q_est_impacto_ativos", v)

            v = st.text_area("Oportunidades climáticas identificadas (§13c)", value=self._q("q_est_oportunidades_desc"),
                             key="_q_est_oport", height=80,
                             placeholder="Ex: Expansão de linha de produtos de baixo carbono, acesso a green bonds...")
            self._q_set("q_est_oportunidades_desc", v)

        # ── 4. Gestão de Riscos ──
        with st.expander("🛡️ 4. Gestão de Riscos (IFRS S2 §23-24)"):
            v = st.text_area(
                "Descreva o processo de identificação e avaliação de riscos climáticos (§23a)",
                value=self._q("q_risk_processo"), key="_q_risk_proc", height=100,
                placeholder="Ex: Análise anual de materialidade com consultoria externa, benchmarking setorial..."
            )
            self._q_set("q_risk_processo", v)

            v = st.selectbox("Frequência de avaliação de riscos climáticos (§23b)",
                             ["Contínua", "Mensal", "Trimestral", "Semestral", "Anual"],
                             index=["Contínua", "Mensal", "Trimestral", "Semestral", "Anual"].index(self._q("q_risk_frequencia")),
                             key="_q_risk_freq")
            self._q_set("q_risk_frequencia", v)

            c1, c2 = st.columns(2)
            with c1:
                v = st.checkbox("Riscos climáticos integrados ao ERM? (§24)", value=self._q("q_risk_integrado"), key="_q_risk_integ")
                self._q_set("q_risk_integrado", v)
            with c2:
                if self._q("q_risk_integrado"):
                    v = st.text_input("Como está integrado?", value=self._q("q_risk_integrado_desc"), key="_q_risk_integ_desc")
                    self._q_set("q_risk_integrado_desc", v)

            v = st.text_area("Principais ações de mitigação em curso (§24b)", value=self._q("q_risk_mitigacao"),
                             key="_q_risk_mit", height=80,
                             placeholder="Ex: Programa de eficiência energética, troca de combustíveis, eletrificação...")
            self._q_set("q_risk_mitigacao", v)

        # ── 5. Metas e Compromissos ──
        with st.expander("🎯 5. Metas Climáticas (IFRS S2 §33-36)"):
            v = st.checkbox("A organização possui metas climáticas formais? (§33)", value=self._q("q_meta_possui"), key="_q_meta_pos")
            self._q_set("q_meta_possui", v)

            if self._q("q_meta_possui"):
                c1, c2 = st.columns(2)
                with c1:
                    v = st.selectbox("Tipo de meta (§34a)", ["Absoluta (tCO₂e)", "Intensidade (tCO₂e/t)", "Ambas"],
                                     index=["Absoluta (tCO₂e)", "Intensidade (tCO₂e/t)", "Ambas"].index(self._q("q_meta_tipo")),
                                     key="_q_meta_tipo")
                    self._q_set("q_meta_tipo", v)
                    v = st.number_input("Ano-base (§34b)", 2015, 2030, self._q("q_meta_base_ano"), key="_q_meta_base")
                    self._q_set("q_meta_base_ano", v)
                with c2:
                    v = st.number_input("Ano-alvo (§34c)", 2025, 2060, self._q("q_meta_alvo_ano"), key="_q_meta_alvo")
                    self._q_set("q_meta_alvo_ano", v)
                    v = st.number_input("Meta de redução (%)", 0, 100, self._q("q_meta_reducao_pct"), key="_q_meta_red")
                    self._q_set("q_meta_reducao_pct", v)

                c1, c2 = st.columns(2)
                with c1:
                    v = st.checkbox("Meta validada por SBTi? (§34e)", value=self._q("q_meta_sbti"), key="_q_meta_sbti")
                    self._q_set("q_meta_sbti", v)
                with c2:
                    v = st.text_input("Ano-alvo Net Zero (se houver)", value=self._q("q_meta_net_zero_ano"), key="_q_meta_nz")
                    self._q_set("q_meta_net_zero_ano", v)

                v = st.text_area("Metas intermediárias / marcos (§35)",
                                 value=self._q("q_meta_intermediarias"), key="_q_meta_interm", height=60,
                                 placeholder="Ex: 2027: -15% Escopo 1&2; 2030: -30% total...")
                self._q_set("q_meta_intermediarias", v)

        # ── 6. Plano de Transição ──
        with st.expander("🔄 6. Plano de Transição (IFRS S2 §14)"):
            v = st.checkbox("A organização possui plano de transição climática? (§14a)",
                            value=self._q("q_trans_possui_plano"), key="_q_trans_plano")
            self._q_set("q_trans_possui_plano", v)

            if self._q("q_trans_possui_plano"):
                v = st.text_area("Principais ações previstas (§14b)", value=self._q("q_trans_acoes"),
                                 key="_q_trans_acoes", height=80,
                                 placeholder="Ex: Substituição de caldeiras a carvão por gás natural até 2028...")
                self._q_set("q_trans_acoes", v)
                v = st.text_input("Investimento previsto (§14c)", value=self._q("q_trans_investimento"),
                                  key="_q_trans_invest", placeholder="Ex: R$ 50 milhões entre 2025-2030")
                self._q_set("q_trans_investimento", v)
                v = st.text_input("Tecnologias planejadas (§14d)", value=self._q("q_trans_tecnologias"),
                                  key="_q_trans_tec", placeholder="Ex: Captura de carbono, hidrogênio verde...")
                self._q_set("q_trans_tecnologias", v)
                v = st.text_area("Dependências e premissas (§14e)", value=self._q("q_trans_dependencias"),
                                 key="_q_trans_dep", height=60,
                                 placeholder="Ex: Disponibilidade de grid renovável, evolução regulatória...")
                self._q_set("q_trans_dependencias", v)

        # ── 7. Verificação ──
        with st.expander("✅ 7. Verificação e Asseguração"):
            c1, c2 = st.columns(2)
            with c1:
                v = st.checkbox("O inventário GEE é assegurado por terceiros?", value=self._q("q_ver_assegurado"), key="_q_ver_ass")
                self._q_set("q_ver_assegurado", v)
            with c2:
                if self._q("q_ver_assegurado"):
                    v = st.selectbox("Tipo de asseguração", ["Limitada", "Razoável"],
                                     index=["Limitada", "Razoável"].index(self._q("q_ver_tipo")), key="_q_ver_tipo")
                    self._q_set("q_ver_tipo", v)
            if self._q("q_ver_assegurado"):
                c1, c2 = st.columns(2)
                with c1:
                    v = st.text_input("Organismo verificador", value=self._q("q_ver_auditor"), key="_q_ver_aud")
                    self._q_set("q_ver_auditor", v)
                with c2:
                    v = st.selectbox("Norma de verificação", ["ISO 14064-3", "ISAE 3410", "AA1000AS", "Outra"],
                                     index=["ISO 14064-3", "ISAE 3410", "AA1000AS", "Outra"].index(self._q("q_ver_norma")),
                                     key="_q_ver_norma")
                    self._q_set("q_ver_norma", v)

        # ── 8. Informações Adicionais ──
        with st.expander("📎 8. Informações Adicionais"):
            c1, c2 = st.columns(2)
            with c1:
                v = st.checkbox("Utiliza compensações de carbono (offsets)?", value=self._q("q_add_offsets"), key="_q_add_off")
                self._q_set("q_add_offsets", v)
                if self._q("q_add_offsets"):
                    v = st.text_input("Detalhes dos offsets", value=self._q("q_add_offsets_desc"), key="_q_add_off_desc")
                    self._q_set("q_add_offsets_desc", v)
            with c2:
                v = st.checkbox("Utiliza preço interno de carbono?", value=self._q("q_add_preco_interno"), key="_q_add_preco")
                self._q_set("q_add_preco_interno", v)
                if self._q("q_add_preco_interno"):
                    v = st.number_input(f"Preço interno (€/{co2e_label(_norm_unit(st.session_state.get('mass_unit', 't')))})", 0, 500, self._q("q_add_preco_valor"), key="_q_add_preco_val")
                    self._q_set("q_add_preco_valor", v)

            v = st.text_area("Notas ou observações adicionais", value=self._q("q_add_notas"),
                             key="_q_add_notas", height=80)
            self._q_set("q_add_notas", v)

        # ── Barra de progresso ──
        st.markdown("---")
        st.markdown("#### 📊 Progresso do Questionário")
        campos_preenchidos = sum(1 for k, v in st.session_state.ifrs_questionario.items()
                                 if v and v != _Q_DEFAULTS.get(k))
        total_campos = len(_Q_DEFAULTS)
        pct = campos_preenchidos / total_campos
        st.progress(pct, text=f"{campos_preenchidos}/{total_campos} campos preenchidos ({pct:.0%})")

        if pct < 0.3:
            st.info("💡 Quanto mais campos preencher, mais completo será seu relatório IFRS.", icon="💡")
        elif pct < 0.7:
            st.info("👍 Bom progresso! Continue preenchendo para um relatório mais robusto.", icon="👍")
        else:
            st.success("🎉 Excelente! Seu relatório terá alta completude.", icon="🎉")

    # ════════════════════════════════════════════════════════════════
    #  INVENTÁRIO GEE – GHG Protocol Corporate Standard
    # ════════════════════════════════════════════════════════════════
    def _render_inventario_ghg(self):
        st.markdown("### 📦 Inventário de Emissões de GEE")
        st.markdown(
            "*Inventário corporativo de gases de efeito estufa conforme o "
            "[GHG Protocol Corporate Accounting and Reporting Standard](https://ghgprotocol.org/corporate-standard). "
            "Escopo 1 = emissões diretas, Escopo 2 = energia indireta, Escopo 3 = cadeia de valor.*"
        )

        unidades = st.session_state.unidades
        edges = st.session_state.edges
        _mu = _norm_unit(st.session_state.get("mass_unit", "t"))
        co2e_lbl = co2e_label(_mu)
        int_lbl = co2e_intensity_label(_mu)
        c2e = lambda v: convert_co2e(v, _mu)
        has_q = self._q_preenchido()
        empresa = self._q("q_empresa_nome") or "Entidade Reportante"
        ano = datetime.now().year
        periodo = (f"{self._q('q_periodo_inicio')} a {self._q('q_periodo_fim')}"
                   if self._q("q_periodo_inicio") else str(ano))

        # ── Cabeçalho do inventário ──
        if has_q:
            c1, c2, c3, c4 = st.columns(4)
            c1.markdown(f"**Entidade:** {empresa}")
            c2.markdown(f"**Período:** {periodo}")
            c3.markdown(f"**Consolidação:** {self._q('q_consolidacao')}")
            c4.markdown(f"**Setor:** {self._q('q_empresa_setor') or '—'}")
        st.markdown("---")

        # ── Construir dados detalhados de fontes ──
        fontes_escopo1, fontes_escopo2, fontes_escopo3 = [], [], []
        resumo_unidades = []

        for u in unidades:
            e1_total = u.IntensidadeEmissaoEscopo1 * u.MassaOutput
            e2_total = u.IntensidadeEmissaoEscopo2 * u.MassaOutput
            e3_total = u.IntensidadeEmissaoEscopo3 * u.MassaOutput
            e_total = e1_total + e2_total + e3_total

            resumo_unidades.append({
                "ID": u.ID_ELO, "Nome": u.Nome, "Localização": u.Localizacao or "—",
                "Escopo 1 (tCO₂e)": round(convert_co2e(e1_total, "t"), 4),
                "Escopo 2 (tCO₂e)": round(convert_co2e(e2_total, "t"), 4),
                "Escopo 3 (tCO₂e)": round(convert_co2e(e3_total, "t"), 4),
                "Total (tCO₂e)": round(convert_co2e(e_total, "t"), 4),
                "Intensidade (tCO₂e/t)": round(convert_co2e(u.IntensidadeEmissao, "t"), 6),
                "Massa Output (t)": round(u.MassaOutput, 2),
            })

            if hasattr(u, "Consumiveis") and u.Consumiveis:
                for i, c in enumerate(u.Consumiveis):
                    if not isinstance(c, dict):
                        continue
                    nome_c = c.get("nome", f"Consumível {i+1}")
                    fator = c.get("fator", 0.0)
                    escopo_raw = str(c.get("escopo", "1")).upper()
                    ce = u.ConsumoEspecifico[i] if i < len(u.ConsumoEspecifico) else 0
                    emissao = fator * ce * u.MassaOutput
                    gas = c.get("gas", "CO₂e")

                    row = {
                        "Unidade": u.ID_ELO,
                        "Nome Unidade": u.Nome,
                        "Localização": u.Localizacao or "—",
                        "Fonte de Emissão": nome_c,
                        "Gás": gas,
                        "Fator de Emissão": round(fator, 6),
                        "Unid. Fator": c.get("unidade_fator", "kgCO₂e/unidade"),
                        "Consumo Específico": round(ce, 6),
                        "Massa Output (t)": round(u.MassaOutput, 2),
                        "Emissão (tCO₂e)": round(convert_co2e(emissao, "t"), 4),
                    }

                    if "1" in escopo_raw:
                        row["Categoria GHG"] = "Combustão estacionária / processos"
                        fontes_escopo1.append(row)
                    elif "2" in escopo_raw:
                        row["Categoria GHG"] = "Eletricidade adquirida / energia"
                        fontes_escopo2.append(row)
                    elif "3" in escopo_raw:
                        row["Categoria GHG"] = "Cadeia de valor (upstream/downstream)"
                        fontes_escopo3.append(row)

        # ── Totais (centralizados em EmissionCalculator — kgCO₂e / t) ──
        _totais = EmissionCalculator.calcular_totais(unidades)
        total_e1 = _totais["escopo1"]
        total_e2 = _totais["escopo2"]
        total_e3 = _totais["escopo3"]
        total_gee = _totais["total"]
        massa_total = _totais["massa_total"]

        # ── KPIs ──
        st.markdown("#### Resumo do Inventário")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("🔴 Escopo 1", f"{c2e(total_e1):,.2f} {co2e_lbl}")
        c2.metric("🟡 Escopo 2", f"{c2e(total_e2):,.2f} {co2e_lbl}")
        c3.metric("🔵 Escopo 3", f"{c2e(total_e3):,.2f} {co2e_lbl}")
        c4.metric("📊 Total GEE", f"{c2e(total_gee):,.2f} {co2e_lbl}")
        c5.metric("📉 Intensidade", f"{total_gee / massa_total:,.4f} {int_lbl}" if massa_total > 0 else "—")

        # ── Tabela‑resumo GHG Protocol ──
        st.markdown("---")
        st.markdown("#### Tabela-Resumo – GHG Protocol Corporate Standard")
        pct = lambda v: f"{v / total_gee * 100:.1f}%" if total_gee > 0 else "0%"
        st.markdown(f"""
| Categoria GHG Protocol | {co2e_lbl} | % do Total | Fontes |
|------------------------|------:|:----------:|-------:|
| **Escopo 1** – Emissões diretas de GEE | {c2e(total_e1):,.4f} | {pct(total_e1)} | {len(fontes_escopo1)} |
| **Escopo 2** – Emissões indiretas de energia | {c2e(total_e2):,.4f} | {pct(total_e2)} | {len(fontes_escopo2)} |
| **Escopo 3** – Outras emissões indiretas | {c2e(total_e3):,.4f} | {pct(total_e3)} | {len(fontes_escopo3)} |
| **Total de Emissões** | **{c2e(total_gee):,.4f}** | **100%** | **{len(fontes_escopo1)+len(fontes_escopo2)+len(fontes_escopo3)}** |
| Escopo 1 + 2 (operacional) | {c2e(total_e1+total_e2):,.4f} | {pct(total_e1+total_e2)} | |
| Intensidade média | {total_gee/massa_total:,.6f} | {int_lbl} | |
""")

        # ── Gráficos ──
        col_pie, col_bar = st.columns(2)
        with col_pie:
            fig_pie = go.Figure(go.Pie(
                labels=["Escopo 1", "Escopo 2", "Escopo 3"],
                values=[c2e(total_e1), c2e(total_e2), c2e(total_e3)],
                marker_colors=[_COLORS["scope1"], _COLORS["scope2"], _COLORS["scope3"]],
                hole=0.45, textinfo="label+percent+value",
                texttemplate="%{label}<br>%{value:,.2f}<br>(%{percent})",
            ))
            fig_pie.update_layout(title="Distribuição por Escopo", height=350,
                                  margin=dict(l=10, r=10, t=40, b=10), showlegend=False)
            st.plotly_chart(fig_pie, use_container_width=True)

        with col_bar:
            # Emissão por localização agrupada por escopo
            loc_data = {}
            for u in unidades:
                loc = u.Localizacao or "Sem local"
                if loc not in loc_data:
                    loc_data[loc] = [0, 0, 0]
                loc_data[loc][0] += u.IntensidadeEmissaoEscopo1 * u.MassaOutput
                loc_data[loc][1] += u.IntensidadeEmissaoEscopo2 * u.MassaOutput
                loc_data[loc][2] += u.IntensidadeEmissaoEscopo3 * u.MassaOutput
            locs_sorted = sorted(loc_data.keys(), key=lambda k: sum(loc_data[k]), reverse=True)
            fig_loc = go.Figure()
            for idx, (lbl, clr) in enumerate([("Escopo 1", _COLORS["scope1"]),
                                               ("Escopo 2", _COLORS["scope2"]),
                                               ("Escopo 3", _COLORS["scope3"])]):
                fig_loc.add_trace(go.Bar(x=locs_sorted, y=[c2e(loc_data[l][idx]) for l in locs_sorted],
                                         name=lbl, marker_color=clr))
            fig_loc.update_layout(barmode="stack", title="Emissões por Localização",
                                  yaxis_title=co2e_lbl, height=350,
                                  margin=dict(l=20, r=20, t=40, b=30), plot_bgcolor="white",
                                  legend=dict(orientation="h", y=1.12, x=0.5, xanchor="center"))
            st.plotly_chart(fig_loc, use_container_width=True)

        # ── Detalhamento por escopo (expansores) ──
        st.markdown("---")
        st.markdown("#### Detalhamento por Fonte de Emissão")

        for scope_n, scope_lbl, scope_clr, scope_rows in [
            (1, "Escopo 1 – Emissões Diretas", _COLORS["scope1"], fontes_escopo1),
            (2, "Escopo 2 – Emissões Indiretas de Energia", _COLORS["scope2"], fontes_escopo2),
            (3, "Escopo 3 – Outras Emissões Indiretas", _COLORS["scope3"], fontes_escopo3),
        ]:
            total_scope = sum(r["Emissão (tCO₂e)"] for r in scope_rows)
            with st.expander(
                f"{'🔴' if scope_n==1 else '🟡' if scope_n==2 else '🔵'} "
                f"{scope_lbl} — {c2e(total_scope):,.4f} {co2e_lbl} ({len(scope_rows)} fontes)",
                expanded=scope_n == 1
            ):
                if scope_n == 1:
                    st.caption("Emissões de fontes controladas pela organização: "
                               "combustão estacionária, processos industriais, veículos próprios. (GHG Protocol Cap. 4)")
                elif scope_n == 2:
                    st.caption("Emissões da geração de eletricidade, calor ou vapor adquiridos e consumidos. "
                               "Método: Baseado em localização. (GHG Protocol Cap. 5)")
                else:
                    st.caption("Emissões indiretas na cadeia de valor: matérias-primas, transporte, "
                               "uso de produtos vendidos, etc. (GHG Protocol Cap. 6 / Scope 3 Standard)")

                if scope_rows:
                    df_scope = pd.DataFrame(scope_rows)
                    df_scope[f"Emissão ({co2e_lbl})"] = df_scope["Emissão (tCO₂e)"].apply(c2e)
                    display_cols = ["Unidade", "Fonte de Emissão", "Gás", "Fator de Emissão",
                                    "Consumo Específico", "Massa Output (t)", f"Emissão ({co2e_lbl})"]
                    st.dataframe(df_scope[display_cols], use_container_width=True, hide_index=True)

                    # Top fontes dentro do escopo
                    top = df_scope.nlargest(5, "Emissão (tCO₂e)")
                    fig_top = go.Figure(go.Bar(
                        y=top["Unidade"] + " — " + top["Fonte de Emissão"],
                        x=top[f"Emissão ({co2e_lbl})"], orientation="h",
                        marker_color=scope_clr, text=top[f"Emissão ({co2e_lbl})"].apply(lambda v: f"{v:,.3f}"),
                        textposition="outside"
                    ))
                    fig_top.update_layout(yaxis=dict(autorange="reversed"), xaxis_title=co2e_lbl,
                                          height=max(200, len(top)*45+80),
                                          margin=dict(l=120, r=60, t=10, b=20), plot_bgcolor="white")
                    st.plotly_chart(fig_top, use_container_width=True)
                else:
                    st.info(f"Nenhuma fonte de Escopo {scope_n} identificada.")

        # ── Resumo por unidade ──
        st.markdown("---")
        st.markdown("#### Resumo por Unidade Produtiva")
        df_resumo = pd.DataFrame(resumo_unidades)
        if not df_resumo.empty:
            df_resumo = df_resumo.sort_values("Total (tCO₂e)", ascending=False)
            _col_map = {
                "Escopo 1 (tCO₂e)": f"Escopo 1 ({co2e_lbl})",
                "Escopo 2 (tCO₂e)": f"Escopo 2 ({co2e_lbl})",
                "Escopo 3 (tCO₂e)": f"Escopo 3 ({co2e_lbl})",
                "Total (tCO₂e)": f"Total ({co2e_lbl})",
                "Intensidade (tCO₂e/t)": f"Intensidade ({int_lbl})",
            }
            df_disp = df_resumo.rename(columns=_col_map)
            for col in [f"Escopo 1 ({co2e_lbl})", f"Escopo 2 ({co2e_lbl})",
                        f"Escopo 3 ({co2e_lbl})", f"Total ({co2e_lbl})"]:
                if col in df_disp.columns:
                    df_disp[col] = df_disp[col].apply(c2e)
            st.dataframe(df_disp, use_container_width=True, hide_index=True)

        # ── Metodologia ──
        with st.expander("📖 Nota Metodológica"):
            st.markdown(f"""
**Padrões aplicados:**
- GHG Protocol – Corporate Accounting and Reporting Standard (Revised Edition)
- GHG Protocol – Scope 3 Standard (Corporate Value Chain)
- ISO 14064-1:2018 – Quantificação de emissões e remoções de GEE

**Abordagem de consolidação:** {self._q('q_consolidacao')}

**Limites organizacionais:** Todas as unidades produtivas cadastradas no modelo, incluindo
operações próprias e fontes identificadas por escopo conforme classificação do GHG Protocol.

**Gases incluídos:** CO₂, CH₄, N₂O e outros conforme fatores de emissão aplicados (expressos em CO₂ equivalente).

**Fatores de emissão:** Valores inseridos pelo usuário conforme base de dados de referência
(IPCC, DEFRA, EPA, Programa Brasileiro GHG Protocol, etc.).

**Exclusões:** Fontes não cadastradas nas unidades produtivas. A completude depende do registro
adequado de todos os consumíveis e fatores de emissão.

**Período de reporte:** {periodo}
""")
            if self._q("q_ver_assegurado"):
                st.markdown(f"**Verificação:** {self._q('q_ver_tipo')} por {self._q('q_ver_auditor') or '—'} "
                            f"conforme {self._q('q_ver_norma')}")

        # ── Downloads ──
        st.markdown("---")
        st.markdown("#### 📥 Exportar Inventário GEE")

        inv_data = {
            "fontes_escopo1": fontes_escopo1,
            "fontes_escopo2": fontes_escopo2,
            "fontes_escopo3": fontes_escopo3,
            "resumo_unidades": resumo_unidades,
            "totais": {"escopo1": convert_co2e(total_e1, "t"),
                       "escopo2": convert_co2e(total_e2, "t"),
                       "escopo3": convert_co2e(total_e3, "t"),
                       "total": convert_co2e(total_gee, "t"),
                       "massa": massa_total},
        }

        # Gerar dados extras para exportação
        inv_data["empresa"] = empresa
        inv_data["periodo"] = periodo
        inv_data["consolidacao"] = self._q("q_consolidacao")

        cfg = get_report_config()
        md_inv = generate_md_inventario(inv_data, cfg)
        pdf_inv = self._gerar_pdf_inventario(inv_data, empresa, periodo) if REPORTLAB_AVAILABLE else None
        fname = f"Inventario_GEE_{empresa.replace(' ','_')}_{ano}"

        extra_dl = {}
        if OPENPYXL_AVAILABLE:
            excel_bytes = self._gerar_excel_inventario(inv_data, empresa, periodo)
            extra_dl["📊 Excel (XLSX)"] = {
                "data": excel_bytes,
                "filename": f"{fname}.xlsx",
                "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        json_inv = json.dumps(inv_data, indent=2, ensure_ascii=False, default=str)
        extra_dl["⬇️ JSON"] = {
            "data": json_inv,
            "filename": f"{fname}.json",
            "mime": "application/json",
        }

        render_download_bar("inventario", md_inv, pdf_inv, extra_dl, fname)

    # ── Geração de Excel do Inventário ──
    def _gerar_excel_inventario(self, inv_data, empresa, periodo):
        wb = Workbook()

        # Estilos
        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        scope1_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        scope2_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        scope3_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        total_fill = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid")
        total_font = Font(name="Calibri", bold=True, size=11)
        title_font = Font(name="Calibri", bold=True, size=14, color="0F766E")
        sub_font = Font(name="Calibri", bold=True, size=12, color="334155")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )
        num_fmt_4 = "0.0000"
        num_fmt_2 = "0.00"

        def _apply_header(ws, row, cols):
            for col_idx, val in enumerate(cols, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

        def _auto_width(ws):
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # ═══ ABA 1: RESUMO ═══
        ws = wb.active
        ws.title = "Resumo"
        ws.sheet_properties.tabColor = "0F766E"

        ws.cell(row=1, column=1, value="INVENTÁRIO DE EMISSÕES DE GEE").font = title_font
        ws.cell(row=2, column=1, value=f"Entidade: {empresa}").font = sub_font
        ws.cell(row=3, column=1, value=f"Período: {periodo}")
        ws.cell(row=4, column=1, value=f"Consolidação: {self._q('q_consolidacao')}")
        ws.cell(row=5, column=1, value=f"Padrão: GHG Protocol Corporate Standard")
        ws.cell(row=6, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        # Tabela resumo
        row = 8
        _apply_header(ws, row, ["Categoria GHG Protocol", "Emissões (tCO₂e)", "% do Total", "Nº Fontes"])
        totals = inv_data["totais"]

        scope_rows = [
            ("Escopo 1 – Emissões diretas", totals["escopo1"], len(inv_data["fontes_escopo1"]), scope1_fill),
            ("Escopo 2 – Emissões indiretas de energia", totals["escopo2"], len(inv_data["fontes_escopo2"]), scope2_fill),
            ("Escopo 3 – Outras emissões indiretas", totals["escopo3"], len(inv_data["fontes_escopo3"]), scope3_fill),
        ]
        for i, (lbl, val, n, fill) in enumerate(scope_rows):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=lbl).fill = fill
            ws.cell(row=r, column=2, value=round(val, 4)).number_format = num_fmt_4
            ws.cell(row=r, column=2).fill = fill
            pct_val = val / totals["total"] * 100 if totals["total"] > 0 else 0
            ws.cell(row=r, column=3, value=f"{pct_val:.1f}%").fill = fill
            ws.cell(row=r, column=4, value=n).fill = fill
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = thin_border

        r_total = row + 4
        for c_idx, val in enumerate(["TOTAL DE EMISSÕES", round(totals["total"], 4), "100%",
                                      sum(len(inv_data[f"fontes_escopo{s}"]) for s in [1,2,3])], 1):
            cell = ws.cell(row=r_total, column=c_idx, value=val)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
        ws.cell(row=r_total, column=2).number_format = num_fmt_4

        r = r_total + 2
        ws.cell(row=r, column=1, value="Escopo 1 + 2 (operacional)").font = Font(bold=True)
        ws.cell(row=r, column=2, value=round(totals["escopo1"] + totals["escopo2"], 4)).number_format = num_fmt_4
        r += 1
        ws.cell(row=r, column=1, value="Intensidade (tCO₂e/t)").font = Font(bold=True)
        ws.cell(row=r, column=2, value=round(totals["total"] / totals["massa"], 6) if totals["massa"] > 0 else 0).number_format = "0.000000"
        r += 1
        ws.cell(row=r, column=1, value="Massa total produzida (t)").font = Font(bold=True)
        ws.cell(row=r, column=2, value=round(totals["massa"], 2)).number_format = num_fmt_2

        _auto_width(ws)

        # ═══ ABA 2-4: Detalhes por escopo ═══
        for scope_n, scope_lbl, fontes, tab_color in [
            (1, "Escopo 1", inv_data["fontes_escopo1"], "EF4444"),
            (2, "Escopo 2", inv_data["fontes_escopo2"], "F59E0B"),
            (3, "Escopo 3", inv_data["fontes_escopo3"], "3B82F6"),
        ]:
            ws_s = wb.create_sheet(title=f"Escopo {scope_n}")
            ws_s.sheet_properties.tabColor = tab_color

            ws_s.cell(row=1, column=1, value=f"ESCOPO {scope_n} – {scope_lbl.upper().split(' – ')[0] if ' – ' in scope_lbl else scope_lbl.upper()}").font = title_font
            ws_s.cell(row=2, column=1, value=empresa)

            if fontes:
                cols = ["Unidade", "Nome Unidade", "Localização", "Fonte de Emissão", "Categoria GHG",
                        "Gás", "Fator de Emissão", "Unid. Fator", "Consumo Específico",
                        "Massa Output (t)", "Emissão (tCO₂e)"]
                _apply_header(ws_s, 4, cols)
                for ri, row_data in enumerate(fontes):
                    for ci, col_name in enumerate(cols):
                        cell = ws_s.cell(row=5 + ri, column=ci + 1, value=row_data.get(col_name, ""))
                        cell.border = thin_border
                        if col_name in ("Fator de Emissão", "Consumo Específico"):
                            cell.number_format = "0.000000"
                        elif col_name == "Emissão (tCO₂e)":
                            cell.number_format = num_fmt_4
                        elif col_name == "Massa Output (t)":
                            cell.number_format = num_fmt_2

                # Total
                r_t = 5 + len(fontes)
                ws_s.cell(row=r_t, column=1, value="TOTAL").font = total_font
                total_scope = sum(f["Emissão (tCO₂e)"] for f in fontes)
                ws_s.cell(row=r_t, column=len(cols), value=round(total_scope, 4)).font = total_font
                ws_s.cell(row=r_t, column=len(cols)).number_format = num_fmt_4
            else:
                ws_s.cell(row=4, column=1, value="Nenhuma fonte identificada para este escopo.")

            _auto_width(ws_s)

        # ═══ ABA 5: UNIDADES ═══
        ws_u = wb.create_sheet(title="Unidades")
        ws_u.sheet_properties.tabColor = "7C3AED"
        ws_u.cell(row=1, column=1, value="RESUMO POR UNIDADE PRODUTIVA").font = title_font

        if inv_data["resumo_unidades"]:
            cols_u = list(inv_data["resumo_unidades"][0].keys())
            _apply_header(ws_u, 3, cols_u)
            for ri, row_data in enumerate(sorted(inv_data["resumo_unidades"],
                                                  key=lambda x: x["Total (tCO₂e)"], reverse=True)):
                for ci, col_name in enumerate(cols_u):
                    cell = ws_u.cell(row=4 + ri, column=ci + 1, value=row_data[col_name])
                    cell.border = thin_border
                    if "tCO₂e" in col_name:
                        cell.number_format = num_fmt_4
                    elif "tCO₂e/t" in col_name:
                        cell.number_format = "0.000000"
                    elif "(t)" in col_name:
                        cell.number_format = num_fmt_2
        _auto_width(ws_u)

        # ═══ ABA 6: METODOLOGIA ═══
        ws_m = wb.create_sheet(title="Metodologia")
        ws_m.sheet_properties.tabColor = "64748B"
        ws_m.cell(row=1, column=1, value="NOTA METODOLÓGICA").font = title_font
        notas = [
            ("Padrão", "GHG Protocol – Corporate Accounting and Reporting Standard (Revised Edition)"),
            ("Escopo 3", "GHG Protocol – Corporate Value Chain (Scope 3) Accounting and Reporting Standard"),
            ("ISO", "ISO 14064-1:2018 – Quantificação de emissões e remoções de GEE"),
            ("Consolidação", self._q("q_consolidacao")),
            ("Limites", "Todas as unidades produtivas cadastradas no modelo"),
            ("Gases", "CO₂, CH₄, N₂O e outros (expresso em CO₂ equivalente)"),
            ("Fatores", "Conforme base de dados de referência (IPCC, DEFRA, EPA, etc.)"),
            ("Período", periodo),
            ("Entidade", empresa),
        ]
        if self._q("q_ver_assegurado"):
            notas.append(("Verificação", f"{self._q('q_ver_tipo')} por {self._q('q_ver_auditor')} ({self._q('q_ver_norma')})"))
        for i, (k, v) in enumerate(notas):
            ws_m.cell(row=3 + i, column=1, value=k).font = Font(bold=True)
            ws_m.cell(row=3 + i, column=2, value=v)
        _auto_width(ws_m)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # ── Geração de PDF do Inventário ──
    def _gerar_pdf_inventario(self, inv_data, empresa, periodo):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2.5*cm, bottomMargin=2*cm)
        story = []
        styles = getSampleStyleSheet()
        q = lambda k: self._q(k)
        t = inv_data["totais"]

        # Estilos
        s_title = ParagraphStyle('IT', parent=styles['Heading1'], fontSize=22,
                                 textColor=rl_colors.HexColor('#0F766E'), spaceAfter=6,
                                 alignment=TA_CENTER, fontName='Helvetica-Bold')
        s_sub = ParagraphStyle('ISub', parent=styles['Normal'], fontSize=11,
                               textColor=rl_colors.HexColor('#64748B'), spaceAfter=24,
                               alignment=TA_CENTER)
        s_h1 = ParagraphStyle('IH1', parent=styles['Heading1'], fontSize=15,
                              textColor=rl_colors.HexColor('#0F766E'), spaceAfter=10,
                              spaceBefore=16, fontName='Helvetica-Bold')
        s_h2 = ParagraphStyle('IH2', parent=styles['Heading2'], fontSize=12,
                              textColor=rl_colors.HexColor('#334155'), spaceAfter=6,
                              spaceBefore=10, fontName='Helvetica-Bold')
        s_body = ParagraphStyle('IB', parent=styles['Normal'], fontSize=9.5,
                                textColor=rl_colors.HexColor('#1E293B'), spaceAfter=6,
                                alignment=TA_JUSTIFY, fontName='Helvetica', leading=13)
        s_note = ParagraphStyle('IN', parent=styles['Normal'], fontSize=8.5,
                                textColor=rl_colors.HexColor('#64748B'), leftIndent=15,
                                rightIndent=15, spaceAfter=8, fontName='Helvetica-Oblique', leading=11)

        def _tbl(data, widths, header_color='#0F766E', alt_color='#F8FAFC'):
            # Sanitize string cells (CO₂ → CO2 for Helvetica)
            clean = [
                [_pdf_safe(c) if isinstance(c, str) else c for c in row]
                for row in data
            ]
            tbl = Table(clean, colWidths=widths)
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor(header_color)),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8.5),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.4, rl_colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor(alt_color)]),
            ]
            tbl.setStyle(TableStyle(style))
            return tbl

        # ═══ CAPA ═══
        story.append(Spacer(1, 3*cm))
        story.append(Paragraph("INVENTÁRIO DE EMISSÕES<br/>DE GASES DE EFEITO ESTUFA", s_title))
        story.append(Paragraph("GHG Protocol – Corporate Accounting and Reporting Standard", s_sub))
        story.append(Spacer(1, 0.5*cm))

        capa = [
            ['DADOS DO INVENTÁRIO', ''],
            ['Entidade', empresa],
            ['CNPJ / Registro', q("q_empresa_cnpj") or "—"],
            ['Setor', q("q_empresa_setor") or "—"],
            ['País', q("q_empresa_pais") or "—"],
            ['Período', periodo],
            ['Consolidação', q("q_consolidacao")],
            ['Responsável', f"{q('q_empresa_responsavel') or '—'} – {q('q_empresa_cargo_responsavel') or ''}"],
        ]
        tbl_capa = _tbl(capa, [5.5*cm, 11*cm])
        tbl_capa.setStyle(TableStyle([
            ('SPAN', (0, 0), (-1, 0)),
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, rl_colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(tbl_capa)
        story.append(Spacer(1, 1*cm))

        # Sumário executivo
        pct_fn = lambda v: f"{v/t['total']*100:.1f}%" if t['total'] > 0 else "0%"
        sum_data = [
            ['RESUMO DE EMISSÕES', '', ''],
            ['Categoria', 'tCO₂e', '% Total'],
            ['Escopo 1 – Emissões diretas', _fmt_num(t['escopo1'], 4), pct_fn(t['escopo1'])],
            ['Escopo 2 – Energia indireta', _fmt_num(t['escopo2'], 4), pct_fn(t['escopo2'])],
            ['Escopo 3 – Cadeia de valor', _fmt_num(t['escopo3'], 4), pct_fn(t['escopo3'])],
            ['Total GEE', _fmt_num(t['total'], 4), '100%'],
            ['Escopo 1+2 (operacional)', _fmt_num(t['escopo1']+t['escopo2'], 4), pct_fn(t['escopo1']+t['escopo2'])],
            ['Intensidade', _fmt_num(t['total']/t['massa'], 6) if t['massa']>0 else '—', 'tCO₂e/t'],
            ['Massa produzida', _fmt_num(t['massa'], 1), 't'],
        ]
        tbl_s = _tbl(sum_data, [7*cm, 5*cm, 4.5*cm])
        tbl_s.setStyle(TableStyle([
            ('SPAN', (0, 0), (-1, 0)),
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, 1), rl_colors.HexColor('#134E4A')),
            ('TEXTCOLOR', (0, 1), (-1, 1), rl_colors.whitesmoke),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('ALIGN', (1, 2), (1, -1), 'RIGHT'),
            ('ALIGN', (2, 2), (2, -1), 'CENTER'),
            ('LINEABOVE', (0, 5), (-1, 5), 1.2, rl_colors.HexColor('#0F766E')),
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, rl_colors.HexColor('#E2E8F0')),
            ('FONTSIZE', (0, 2), (-1, -1), 9),
        ]))
        story.append(tbl_s)
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(
            f"<i>Inventário elaborado conforme GHG Protocol Corporate Standard, ISO 14064-1. "
            f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}.</i>", s_note))
        story.append(PageBreak())

        # ═══ DETALHAMENTO POR ESCOPO ═══
        for scope_n, scope_lbl, scope_clr, fontes in [
            (1, "ESCOPO 1 – EMISSÕES DIRETAS", '#EF4444', inv_data["fontes_escopo1"]),
            (2, "ESCOPO 2 – EMISSÕES INDIRETAS DE ENERGIA", '#F59E0B', inv_data["fontes_escopo2"]),
            (3, "ESCOPO 3 – OUTRAS EMISSÕES INDIRETAS", '#3B82F6', inv_data["fontes_escopo3"]),
        ]:
            story.append(Paragraph(f"{scope_lbl}", s_h1))

            if scope_n == 1:
                story.append(Paragraph(
                    "Emissões de fontes possuídas ou controladas pela organização: combustão estacionária "
                    "e móvel, processos industriais, emissões fugitivas. (GHG Protocol, Capítulo 4)", s_body))
            elif scope_n == 2:
                story.append(Paragraph(
                    "Emissões associadas à geração de eletricidade, calor ou vapor adquiridos e consumidos. "
                    "Método: baseado em localização. (GHG Protocol, Capítulo 5)", s_body))
            else:
                story.append(Paragraph(
                    "Emissões indiretas na cadeia de valor: compra de bens e serviços, transporte e distribuição, "
                    "uso de produtos vendidos, investimentos, etc. (GHG Protocol, Capítulo 6 / Scope 3 Standard)", s_body))

            if fontes:
                total_scope = sum(f["Emissão (tCO₂e)"] for f in fontes)
                story.append(Paragraph(_pdf_safe(f"<b>Total Escopo {scope_n}: {_fmt_num(total_scope, 4)} tCO₂e "
                                        f"({len(fontes)} fontes)</b>"), s_body))
                story.append(Spacer(1, 0.2*cm))

                # Tabela de fontes
                hdr = ['Unidade', 'Fonte', 'Gás', 'FE', 'CE', 'Massa (t)', 'tCO₂e']
                rows_tbl = [hdr]
                for f in sorted(fontes, key=lambda x: x["Emissão (tCO₂e)"], reverse=True):
                    rows_tbl.append([
                        f["Unidade"],
                        Paragraph(f["Fonte de Emissão"][:30], s_body),
                        f.get("Gás", "CO₂e"),
                        _fmt_num(f['Fator de Emissão'], 4),
                        _fmt_num(f['Consumo Específico'], 4),
                        _fmt_num(f['Massa Output (t)'], 1),
                        _fmt_num(f['Emissão (tCO₂e)'], 4),
                    ])
                rows_tbl.append(['TOTAL', '', '', '', '', '', _fmt_num(total_scope, 4)])
                tbl_f = _tbl(rows_tbl, [2.2*cm, 4*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 2.8*cm], scope_clr)
                tbl_f.setStyle(TableStyle([
                    ('FONTNAME', (0, len(rows_tbl)-1), (-1, len(rows_tbl)-1), 'Helvetica-Bold'),
                    ('LINEABOVE', (0, len(rows_tbl)-1), (-1, len(rows_tbl)-1), 1, rl_colors.HexColor(scope_clr)),
                    ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ]))
                story.append(tbl_f)
            else:
                story.append(Paragraph(f"<i>Nenhuma fonte de Escopo {scope_n} identificada.</i>", s_note))

            story.append(Spacer(1, 0.5*cm))

        story.append(PageBreak())

        # ═══ RESUMO POR UNIDADE ═══
        story.append(Paragraph("RESUMO POR UNIDADE PRODUTIVA", s_h1))
        sorted_units = sorted(inv_data["resumo_unidades"], key=lambda x: x["Total (tCO₂e)"], reverse=True)[:20]
        u_hdr = ['ID', 'Nome', 'Local', 'E1', 'E2', 'E3', 'Total', 'Intens.']
        u_rows = [u_hdr]
        for u in sorted_units:
            u_rows.append([
                u["ID"], Paragraph(str(u["Nome"])[:22], s_body),
                str(u["Localização"])[:12],
                _fmt_num(u['Escopo 1 (tCO₂e)'], 2),
                _fmt_num(u['Escopo 2 (tCO₂e)'], 2),
                _fmt_num(u['Escopo 3 (tCO₂e)'], 2),
                _fmt_num(u['Total (tCO₂e)'], 2),
                _fmt_num(u['Intensidade (tCO₂e/t)'], 4),
            ])
        tbl_u = _tbl(u_rows, [2*cm, 3.2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2.3*cm], '#334155')
        tbl_u.setStyle(TableStyle([('ALIGN', (3, 1), (-1, -1), 'RIGHT')]))
        story.append(tbl_u)
        story.append(Paragraph(_pdf_safe(
            f"<i>Top {len(sorted_units)} unidades por emissão total (tCO₂e). "
            f"Total de unidades: {len(inv_data['resumo_unidades'])}.</i>"), s_note))

        # ═══ METODOLOGIA ═══
        story.append(Spacer(1, 0.8*cm))
        story.append(Paragraph("NOTA METODOLÓGICA", s_h1))
        met_items = [
            "GHG Protocol – Corporate Accounting and Reporting Standard (Revised Edition, 2015)",
            "GHG Protocol – Corporate Value Chain (Scope 3) Accounting and Reporting Standard",
            "ISO 14064-1:2018 – Quantificação e reporte de emissões e remoções de GEE",
            f"Abordagem de consolidação: {self._q('q_consolidacao')}",
            _pdf_safe("Gases incluídos: CO₂, CH₄, N₂O e outros (expressos em CO₂ equivalente – GWP AR5/AR6)"),
            "Fatores de emissão: conforme bases IPCC, DEFRA, EPA, Programa Brasileiro GHG Protocol",
            f"Período: {periodo}",
        ]
        for m in met_items:
            story.append(Paragraph(f"• {m}", s_body))

        if self._q("q_ver_assegurado"):
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph(
                f"<b>Verificação:</b> Asseguração {self._q('q_ver_tipo').lower()} realizada por "
                f"{self._q('q_ver_auditor') or '—'} conforme {self._q('q_ver_norma')}.", s_body))

        # Disclaimer
        story.append(Spacer(1, 0.8*cm))
        story.append(HRFlowable(width="100%", color=rl_colors.HexColor('#E2E8F0')))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            "<b>Disclaimer:</b> Este inventário foi gerado automaticamente com base nos dados modelados "
            "na Calculadora de Pegada de Carbono de Materiais. A completude e exatidão dependem do "
            "registro adequado de todas as fontes de emissão, fatores de emissão e dados de atividade. "
            "Recomenda-se verificação por terceiros independentes conforme ISO 14064-3 ou ISAE 3410 "
            "antes da publicação oficial.", s_note))
        if self._q("q_add_notas"):
            story.append(Paragraph(f"<b>Observações:</b> {self._q('q_add_notas')}", s_note))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            f"<i>Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
            f"Padrões: GHG Protocol, ISO 14064-1, IFRS S2 | Calculadora CMP v2.2</i>", s_note))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    # ════════════════════════════════════════════════════════════════
    #  REPORTE IFRS S1/S2 (WEB) – COMPLETO COM QUESTIONÁRIO
    # ════════════════════════════════════════════════════════════════
    def _render_ifrs_report(self):
        has_q = self._q_preenchido()
        empresa = self._q("q_empresa_nome") or "Entidade Reportante"

        if not has_q:
            st.warning("⚠️ Preencha o **Questionário IFRS** (aba anterior) para gerar um relatório completo e personalizado.", icon="⚠️")

        st.markdown(f"### 📋 Relatório IFRS S1/S2 – Divulgações Climáticas")
        if has_q:
            periodo = f"{self._q('q_periodo_inicio')} a {self._q('q_periodo_fim')}" if self._q("q_periodo_inicio") else str(datetime.now().year)
            st.markdown(f"**{empresa}** · Período: {periodo} · Consolidação: {self._q('q_consolidacao')}")

        unidades = st.session_state.unidades
        edges = st.session_state.edges
        ano_reporte = datetime.now().year

        # Cálculos
        esc1 = sum(u.IntensidadeEmissaoEscopo1 * u.MassaOutput for u in unidades)
        esc2 = sum(u.IntensidadeEmissaoEscopo2 * u.MassaOutput for u in unidades)
        esc3 = sum(u.IntensidadeEmissaoEscopo3 * u.MassaOutput for u in unidades)
        total_ghg = esc1 + esc2 + esc3
        massa_total = sum(u.MassaOutput for u in unidades)
        intensidade_media = total_ghg / massa_total if massa_total > 0 else 0

        unidades_tf = [u for u in unidades if u.TaxacaoFronteira]
        unidades_tl = [u for u in unidades if u.TaxacaoLocal]
        em_tf = sum(u.IntensidadeEmissao * u.MassaOutput for u in unidades_tf)
        em_tl = sum(u.IntensidadeEmissao * u.MassaOutput for u in unidades_tl)
        pct_taxada = (em_tf + em_tl) / total_ghg * 100 if total_ghg > 0 else 0

        high_int = [u for u in unidades if u.IntensidadeEmissao > intensidade_media * 2]
        low_int = [u for u in unidades if 0 < u.IntensidadeEmissao < intensidade_media * 0.5]

        # ── Unidades de exibição ──
        _mu = _norm_unit(st.session_state.get("mass_unit", "t"))
        co2e_lbl = co2e_label(_mu)
        int_lbl = co2e_intensity_label(_mu)
        c2e = lambda v: convert_co2e(v, _mu)
        c2i = lambda v: convert_co2e(v, "t")  # intensidade: sempre ÷1000

        # ── 1. GOVERNANÇA ──
        with st.expander("🏛️ 1. Governança (IFRS S2 §5-12)", expanded=True):
            if has_q and self._q("q_gov_orgao"):
                st.markdown(f"**Órgão de supervisão:** {self._q('q_gov_orgao')}")
                st.markdown(f"**Frequência de reporte ao órgão:** {self._q('q_gov_frequencia')}")
                if self._q("q_gov_comite"):
                    st.markdown(f"**Comitê dedicado:** {self._q('q_gov_comite_nome') or 'Sim'}")
                if self._q("q_gov_competencias"):
                    st.markdown(f"**Competências climáticas:** {self._q('q_gov_competencias')}")
                if self._q("q_gov_integracao_estrategia"):
                    st.markdown(f"**Integração na estratégia:** {self._q('q_gov_integracao_estrategia')}")
                if self._q("q_gov_remuneracao_vinculada"):
                    st.markdown(f"**Remuneração vinculada a metas climáticas:** Sim – {self._q('q_gov_remuneracao_detalhes')}")
            else:
                st.markdown(
                    "> *A entidade deve divulgar o(s) órgão(s) de governança responsáveis pela supervisão "
                    "dos riscos climáticos (IFRS S2 §6). Preencha o questionário para detalhar.*"
                )

        # ── 2. ESTRATÉGIA ──
        with st.expander("📐 2. Estratégia (IFRS S2 §13-22)", expanded=True):
            # Riscos automáticos + questionário
            st.markdown("#### 2.1 Riscos Identificados")

            riscos = []
            # Riscos automáticos (dos dados)
            if unidades_tf or unidades_tl:
                riscos.append({
                    "tipo": "Transição – Regulatório",
                    "descricao": f"Taxação de carbono: {len(unidades_tf)} unidade(s) sob CBAM e {len(unidades_tl)} sob taxação local.",
                    "exposicao": f"{c2e(em_tf + em_tl):,.2f} {co2e_lbl} ({pct_taxada:.1f}%)",
                    "horizonte": self._q("q_est_horizonte_curto") or "Curto prazo",
                })
            if unidades:
                sorted_u = sorted(unidades, key=lambda x: x.IntensidadeEmissao * x.MassaOutput, reverse=True)
                top1 = sorted_u[0]
                top1_pct = (top1.IntensidadeEmissao * top1.MassaOutput) / total_ghg * 100 if total_ghg > 0 else 0
                if top1_pct > 40:
                    riscos.append({
                        "tipo": "Transição – Concentração",
                        "descricao": f"Unidade {top1.ID_ELO} concentra {top1_pct:.1f}% das emissões totais.",
                        "exposicao": f"{c2e(top1.IntensidadeEmissao * top1.MassaOutput):,.2f} {co2e_lbl}",
                        "horizonte": self._q("q_est_horizonte_medio") or "Médio prazo",
                    })
            if high_int:
                riscos.append({
                    "tipo": "Transição – Tecnológico",
                    "descricao": f"{len(high_int)} unidade(s) com intensidade >2× a média ({c2i(intensidade_media):.4f} {int_lbl}).",
                    "exposicao": f"{c2e(sum(u.IntensidadeEmissao * u.MassaOutput for u in high_int)):,.2f} {co2e_lbl}",
                    "horizonte": self._q("q_est_horizonte_medio") or "Médio prazo",
                })

            # Riscos do questionário
            for rf in self._q("q_est_riscos_fisicos"):
                riscos.append({"tipo": "Físico", "descricao": rf, "exposicao": "A avaliar", "horizonte": self._q("q_est_horizonte_longo") or "Longo prazo"})
            for rt in self._q("q_est_riscos_transicao"):
                if "Regulação" not in rt or not (unidades_tf or unidades_tl):  # evita duplicar
                    riscos.append({"tipo": "Transição", "descricao": rt, "exposicao": "A avaliar", "horizonte": self._q("q_est_horizonte_medio") or "Médio prazo"})

            if riscos:
                df_riscos = pd.DataFrame(riscos)
                st.dataframe(df_riscos, use_container_width=True, hide_index=True)
            else:
                st.success("Nenhum risco climático material identificado.")

            # Oportunidades
            st.markdown("#### 2.2 Oportunidades Identificadas")
            oportunidades = []
            if low_int:
                red_pot = sum(u.IntensidadeEmissao * u.MassaOutput for u in high_int) if high_int else 0
                oportunidades.append({
                    "tipo": "Eficiência de Recursos",
                    "descricao": f"{len(low_int)} unidade(s) benchmark (<50% média). Potencial de replicação para reduzir {c2e(red_pot):,.2f} {co2e_lbl}.",
                })
            untaxed = [u for u in unidades if not u.TaxacaoFronteira and not u.TaxacaoLocal]
            if untaxed and (unidades_tf or unidades_tl):
                oportunidades.append({
                    "tipo": "Resiliência Regulatória",
                    "descricao": f"{len(untaxed)} unidade(s) fora de jurisdições de taxação – vantagem competitiva.",
                })
            if self._q("q_est_oportunidades_desc"):
                oportunidades.append({"tipo": "Identificada pela organização", "descricao": self._q("q_est_oportunidades_desc")})

            if oportunidades:
                for o in oportunidades:
                    st.markdown(f"- **{o['tipo']}:** {o['descricao']}")
            else:
                st.caption("Nenhuma oportunidade identificada.")

            # Impacto financeiro
            st.markdown("#### 2.3 Impacto Financeiro (§21)")
            if has_q and any(self._q(f"q_est_impacto_{k}") for k in ["receita", "custos", "ativos"]):
                c1, c2, c3 = st.columns(3)
                if self._q("q_est_impacto_receita"):
                    c1.info(f"📉 **Receitas:** {self._q('q_est_impacto_receita')}")
                if self._q("q_est_impacto_custos"):
                    c2.warning(f"💰 **Custos:** {self._q('q_est_impacto_custos')}")
                if self._q("q_est_impacto_ativos"):
                    c3.error(f"🏢 **Ativos:** {self._q('q_est_impacto_ativos')}")

            st.markdown("##### 💰 Simulação de Custo de Carbono")
            default_price = self._q("q_add_preco_valor") if self._q("q_add_preco_interno") else 50
            preco_co2 = st.slider(f"Preço do carbono (€/{co2e_lbl})", 0, 200, max(0, default_price), 5, key="preco_co2_v3")
            moeda = self._q("q_moeda").split(" ")[0] if has_q else "EUR"
            simbolo = {"BRL": "R$", "USD": "$", "EUR": "€", "GBP": "£"}.get(moeda, "€")
            c1, c2, c3 = st.columns(3)
            c1.metric(f"Custo Total", f"{simbolo} {c2e(total_ghg) * preco_co2:,.0f}")
            c2.metric(f"Custo CBAM/Fronteira", f"{simbolo} {c2e(em_tf) * preco_co2:,.0f}")
            c3.metric(f"Custo Local", f"{simbolo} {c2e(em_tl) * preco_co2:,.0f}")

            if self._q("q_est_cenarios"):
                st.markdown(f"**Cenários climáticos utilizados (§22):** {self._q('q_est_cenarios_detalhes')}")

        # ── 3. GESTÃO DE RISCOS ──
        with st.expander("🛡️ 3. Gestão de Riscos (IFRS S2 §23-24)", expanded=False):
            if has_q and self._q("q_risk_processo"):
                st.markdown(f"**Processo de identificação:** {self._q('q_risk_processo')}")
                st.markdown(f"**Frequência de avaliação:** {self._q('q_risk_frequencia')}")
                if self._q("q_risk_integrado"):
                    st.markdown(f"**Integração com ERM:** Sim – {self._q('q_risk_integrado_desc')}")
                if self._q("q_risk_mitigacao"):
                    st.markdown(f"**Ações de mitigação:** {self._q('q_risk_mitigacao')}")
            else:
                st.markdown("**Critérios de materialidade (automatizados):**\n"
                            "- Intensidade de emissão >2× a média do portfólio\n"
                            "- Concentração >40% em uma única unidade\n"
                            "- Exposição a taxação de carbono")

            st.markdown("#### Mapa de Risco das Unidades")
            self._render_risk_matrix(unidades, intensidade_media)

        # ── 4. MÉTRICAS E METAS ──
        with st.expander("📊 4. Métricas e Metas (IFRS S2 §29-36)", expanded=True):
            st.markdown("#### 4.1 Inventário de GEE (§29a)")
            st.markdown(f"""
| Métrica | Valor | Unidade | Padrão |
|---------|------:|---------|--------|
| **Escopo 1** – Emissões diretas | {c2e(esc1):,.2f} | {co2e_lbl} | GHG Protocol |
| **Escopo 2** – Energia (localização) | {c2e(esc2):,.2f} | {co2e_lbl} | GHG Protocol |
| **Escopo 3** – Cadeia de valor | {c2e(esc3):,.2f} | {co2e_lbl} | GHG Protocol |
| **Total GEE** | **{c2e(total_ghg):,.2f}** | **{co2e_lbl}** | |
| Intensidade de emissão | {c2i(intensidade_media):,.4f} | {int_lbl} produto | IFRS S2 §29(d) |
| Massa total produzida | {massa_total:,.1f} | t | |
| Abordagem de consolidação | {self._q('q_consolidacao')} | | IFRS S1 §B25 |
""")

            st.markdown("#### 4.2 Exposição Regulatória (§29b)")
            st.markdown(f"""
| Métrica | Valor | Ref. |
|---------|------:|------|
| GEE sob regulação de preço | {pct_taxada:.1f}% | §29(b) |
| Emissões sob CBAM/fronteira | {c2e(em_tf):,.2f} {co2e_lbl} | §29(b) |
| Emissões sob taxação local | {c2e(em_tl):,.2f} {co2e_lbl} | §29(b) |
| Unidades de alto risco | {len(high_int)} | §29(f) |
""")

            # Waterfall
            fig_wf = go.Figure(go.Waterfall(
                orientation="v",
                measure=["relative", "relative", "relative", "total"],
                x=["Escopo 1", "Escopo 2", "Escopo 3", "Total GEE"],
                y=[c2e(esc1), c2e(esc2), c2e(esc3), 0],
                connector={"line": {"color": _COLORS["muted"]}},
                decreasing={"marker": {"color": _COLORS["scope3"]}},
                increasing={"marker": {"color": _COLORS["scope1"]}},
                totals={"marker": {"color": _COLORS["primary"]}},
                text=[f"{c2e(esc1):,.1f}", f"{c2e(esc2):,.1f}", f"{c2e(esc3):,.1f}", f"{c2e(total_ghg):,.1f}"],
                textposition="outside",
            ))
            fig_wf.update_layout(yaxis_title=co2e_lbl, height=320, margin=dict(l=20, r=20, t=30, b=20), plot_bgcolor="white")
            st.plotly_chart(fig_wf, use_container_width=True)

            # Metas
            st.markdown("#### 4.3 Metas Climáticas (§33-36)")
            if self._q("q_meta_possui"):
                meta_red = self._q("q_meta_reducao_pct")
                ano_base = self._q("q_meta_base_ano")
                ano_alvo = self._q("q_meta_alvo_ano")
                emissao_meta = total_ghg * (1 - meta_red / 100)

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Tipo", self._q("q_meta_tipo"))
                c2.metric("Ano-base", str(ano_base))
                c3.metric("Ano-alvo", str(ano_alvo))
                c4.metric("Redução", f"{meta_red}%")

                c1, c2, c3 = st.columns(3)
                c1.metric("Emissão Atual", f"{c2e(total_ghg):,.2f} {co2e_lbl}")
                c2.metric("Emissão-Meta", f"{c2e(emissao_meta):,.2f} {co2e_lbl}")
                c3.metric("Validação SBTi", "✅ Sim" if self._q("q_meta_sbti") else "❌ Não")

                if self._q("q_meta_net_zero_ano"):
                    st.markdown(f"**Compromisso Net Zero:** {self._q('q_meta_net_zero_ano')}")
                if self._q("q_meta_intermediarias"):
                    st.markdown(f"**Marcos intermediários:** {self._q('q_meta_intermediarias')}")

                # Trajetória
                anos = list(range(ano_base, int(ano_alvo) + 1))
                n = len(anos)
                traj = [c2e(total_ghg) - (c2e(total_ghg) - c2e(emissao_meta)) * i / (n - 1) for i in range(n)] if n > 1 else [c2e(total_ghg)]
                fig_t = go.Figure()
                fig_t.add_trace(go.Scatter(x=anos, y=traj, mode="lines+markers", name="Trajetória",
                                           line=dict(color=_COLORS["primary"], width=3), marker=dict(size=7)))
                fig_t.add_trace(go.Scatter(x=[ano_base], y=[c2e(total_ghg)], mode="markers", name="Atual",
                                           marker=dict(color=_COLORS["danger"], size=14, symbol="diamond")))
                fig_t.add_hline(y=c2e(emissao_meta), line_dash="dash", line_color=_COLORS["success"],
                                annotation_text=f"Meta {ano_alvo}: {c2e(emissao_meta):,.0f}")
                fig_t.update_layout(yaxis_title=co2e_lbl, xaxis_title="Ano", height=320,
                                    margin=dict(l=20, r=20, t=30, b=30), plot_bgcolor="white",
                                    legend=dict(orientation="h", y=1.1, x=0.5, xanchor="center"))
                st.plotly_chart(fig_t, use_container_width=True)
            else:
                meta_red = st.number_input("Meta de redução (%)", 0, 100, 30, key="meta_red_fallback")
                ano_alvo = st.number_input("Ano-alvo", 2025, 2050, 2030, key="ano_alvo_fallback")
                emissao_meta = total_ghg * (1 - meta_red / 100)
                st.caption("💡 Defina metas formais no questionário para um relatório mais completo.")

        # ── 5. PLANO DE TRANSIÇÃO ──
        with st.expander("🔄 5. Plano de Transição (IFRS S2 §14)", expanded=False):
            if self._q("q_trans_possui_plano"):
                if self._q("q_trans_acoes"):
                    st.markdown(f"**Ações previstas:** {self._q('q_trans_acoes')}")
                if self._q("q_trans_investimento"):
                    st.markdown(f"**Investimento previsto:** {self._q('q_trans_investimento')}")
                if self._q("q_trans_tecnologias"):
                    st.markdown(f"**Tecnologias planejadas:** {self._q('q_trans_tecnologias')}")
                if self._q("q_trans_dependencias"):
                    st.markdown(f"**Dependências e premissas:** {self._q('q_trans_dependencias')}")
            else:
                st.caption("Nenhum plano de transição registrado. Defina no questionário (seção 6).")

            tec_set = set()
            for u in unidades:
                if hasattr(u, "Tecnologia") and u.Tecnologia:
                    tec_set.add(u.Tecnologia.nome)
            if tec_set:
                st.markdown(f"**Tecnologias atuais:** {', '.join(sorted(tec_set))}")

        # ── 6. VERIFICAÇÃO ──
        with st.expander("✅ 6. Verificação e Asseguração", expanded=False):
            if self._q("q_ver_assegurado"):
                st.markdown(f"**Asseguração:** {self._q('q_ver_tipo')}")
                st.markdown(f"**Auditor:** {self._q('q_ver_auditor')}")
                st.markdown(f"**Norma:** {self._q('q_ver_norma')}")
            else:
                st.caption("Inventário GEE não assegurado por terceiros. Considere para conformidade completa com IFRS S2.")

        # ── 7. IFRS S1 ──
        with st.expander("📜 7. IFRS S1 – Requisitos Gerais", expanded=False):
            checks = {
                "Identificação da entidade": has_q,
                "Governança": bool(self._q("q_gov_orgao")),
                "Estratégia (riscos/oportunidades)": len(riscos) > 0,
                "Gestão de Riscos": bool(self._q("q_risk_processo")),
                "Métricas GEE (Escopos 1-3)": total_ghg > 0,
                "Metas climáticas": self._q("q_meta_possui"),
                "Plano de transição": self._q("q_trans_possui_plano"),
                "Cadeia de valor mapeada": len(edges) > 0,
                "Verificação/Asseguração": self._q("q_ver_assegurado"),
                "Análise de cenários": self._q("q_est_cenarios"),
            }
            for item, ok in checks.items():
                st.markdown(f"{'✅' if ok else '⬜'} {item}")
            completude = sum(checks.values()) / len(checks)
            st.progress(completude, text=f"Conformidade IFRS S1/S2: {completude:.0%}")

        # ── EXPORTAÇÃO ──
        st.markdown("---")
        st.markdown("#### 📥 Exportar Relatório")

        reporte_data = self._build_reporte_data(
            ano_reporte, esc1, esc2, esc3, total_ghg, massa_total, intensidade_media,
            em_tf, em_tl, pct_taxada, riscos, oportunidades, unidades, empresa
        )

        cfg = get_report_config()
        md_ifrs = generate_md_ifrs(reporte_data, cfg)
        pdf_ifrs = self._gerar_pdf_ifrs_s2(reporte_data, riscos, oportunidades, high_int, unidades) if REPORTLAB_AVAILABLE else None
        fname = f"IFRS_S2_{empresa.replace(' ','_')}_{ano_reporte}"

        extra_dl = {
            "⬇️ JSON": {
                "data": json.dumps(reporte_data, indent=2, ensure_ascii=False, default=str),
                "filename": f"{fname}.json",
                "mime": "application/json",
            },
        }
        df_exp = pd.DataFrame(reporte_data.get("detalhamento_unidades", []))
        if not df_exp.empty:
            extra_dl["⬇️ CSV"] = {
                "data": df_exp.to_csv(index=False),
                "filename": f"unidades_{ano_reporte}.csv",
                "mime": "text/csv",
            }

        render_download_bar("ifrs", md_ifrs, pdf_ifrs, extra_dl, fname)

    # ── Construir dicionário do reporte ──
    def _build_reporte_data(self, ano, e1, e2, e3, total, massa, intens, em_tf, em_tl, pct_t,
                             riscos, oportunidades, unidades, empresa):
        q = lambda k: self._q(k)
        data = {
            "padrao": "IFRS S2 – Climate-Related Disclosures",
            "entidade": {
                "nome": empresa,
                "cnpj": q("q_empresa_cnpj"),
                "setor": q("q_empresa_setor"),
                "pais": q("q_empresa_pais"),
                "responsavel": q("q_empresa_responsavel"),
                "cargo": q("q_empresa_cargo_responsavel"),
                "consolidacao": q("q_consolidacao"),
            },
            "periodo": {"inicio": q("q_periodo_inicio"), "fim": q("q_periodo_fim"), "ano": ano},
            "governanca": {
                "orgao_supervisor": q("q_gov_orgao"),
                "frequencia": q("q_gov_frequencia"),
                "comite_dedicado": q("q_gov_comite"),
                "comite_nome": q("q_gov_comite_nome"),
                "competencias": q("q_gov_competencias"),
                "integracao_estrategia": q("q_gov_integracao_estrategia"),
                "remuneracao_vinculada": q("q_gov_remuneracao_vinculada"),
            },
            "metricas_ghg": {
                "escopo_1_tco2e": round(e1, 4), "escopo_2_tco2e": round(e2, 4),
                "escopo_3_tco2e": round(e3, 4), "total_tco2e": round(total, 4),
                "intensidade_media": round(intens, 6), "massa_total_t": round(massa, 2),
            },
            "exposicao_regulatoria": {
                "emissao_cbam_tco2e": round(em_tf, 4), "emissao_local_tco2e": round(em_tl, 4),
                "pct_sob_regulacao": round(pct_t, 2),
            },
            "riscos": riscos, "oportunidades": oportunidades,
            "metas": {
                "possui": q("q_meta_possui"), "tipo": q("q_meta_tipo"),
                "ano_base": q("q_meta_base_ano"), "ano_alvo": q("q_meta_alvo_ano"),
                "reducao_pct": q("q_meta_reducao_pct"), "sbti": q("q_meta_sbti"),
                "net_zero": q("q_meta_net_zero_ano"),
            },
            "plano_transicao": {
                "possui": q("q_trans_possui_plano"), "acoes": q("q_trans_acoes"),
                "investimento": q("q_trans_investimento"),
            },
            "verificacao": {
                "assegurado": q("q_ver_assegurado"), "tipo": q("q_ver_tipo"),
                "auditor": q("q_ver_auditor"), "norma": q("q_ver_norma"),
            },
            "detalhamento_unidades": [{
                "id": u.ID_ELO, "nome": u.Nome, "localizacao": u.Localizacao,
                "escopo1": round(u.IntensidadeEmissaoEscopo1 * u.MassaOutput, 4),
                "escopo2": round(u.IntensidadeEmissaoEscopo2 * u.MassaOutput, 4),
                "escopo3": round(u.IntensidadeEmissaoEscopo3 * u.MassaOutput, 4),
                "total": round(u.IntensidadeEmissao * u.MassaOutput, 4),
                "intensidade": round(u.IntensidadeEmissao, 6),
                "massa_output": round(u.MassaOutput, 2),
                "taxacao_fronteira": u.TaxacaoFronteira, "taxacao_local": u.TaxacaoLocal,
            } for u in unidades],
        }
        return data

    # ════════════════════════════════════════════════════════════════
    #  GERAÇÃO DE PDF – COMPLETA COM QUESTIONÁRIO
    # ════════════════════════════════════════════════════════════════
    def _gerar_pdf_ifrs_s2(self, reporte, riscos, oportunidades, high_intensity, unidades):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2.5*cm, bottomMargin=2*cm)
        story = []
        styles = getSampleStyleSheet()
        q = lambda k: self._q(k)
        rd = reporte  # shortcut

        # Estilos
        s_title = ParagraphStyle('T', parent=styles['Heading1'], fontSize=22,
                                 textColor=rl_colors.HexColor('#0F766E'), spaceAfter=6,
                                 alignment=TA_CENTER, fontName='Helvetica-Bold')
        s_sub = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=11,
                               textColor=rl_colors.HexColor('#64748B'), spaceAfter=30,
                               alignment=TA_CENTER, fontName='Helvetica')
        s_h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=15,
                              textColor=rl_colors.HexColor('#0F766E'), spaceAfter=10,
                              spaceBefore=18, fontName='Helvetica-Bold')
        s_h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=12,
                              textColor=rl_colors.HexColor('#334155'), spaceAfter=6,
                              spaceBefore=12, fontName='Helvetica-Bold')
        s_body = ParagraphStyle('B', parent=styles['Normal'], fontSize=9.5,
                                textColor=rl_colors.HexColor('#1E293B'), spaceAfter=6,
                                alignment=TA_JUSTIFY, fontName='Helvetica', leading=13)
        s_note = ParagraphStyle('N', parent=styles['Normal'], fontSize=8.5,
                                textColor=rl_colors.HexColor('#64748B'), leftIndent=15,
                                rightIndent=15, spaceAfter=8, fontName='Helvetica-Oblique', leading=11)
        s_bullet = ParagraphStyle('Bul', parent=s_body, leftIndent=15, bulletIndent=5)

        def _tbl(data, widths, header_color='#0F766E', alt_color='#F8FAFC'):
            # Sanitize string cells (CO₂ → CO2 for Helvetica)
            clean = [
                [_pdf_safe(c) if isinstance(c, str) else c for c in row]
                for row in data
            ]
            t = Table(clean, colWidths=widths)
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor(header_color)),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8.5),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.4, rl_colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor(alt_color)]),
            ]
            t.setStyle(TableStyle(style))
            return t

        empresa = rd["entidade"]["nome"] or "Entidade"
        periodo = f"{q('q_periodo_inicio')} a {q('q_periodo_fim')}" if q("q_periodo_inicio") else str(rd["periodo"]["ano"])
        ghg = rd["metricas_ghg"]

        # ═══ CAPA ═══
        story.append(Spacer(1, 2.5*cm))
        story.append(Paragraph("RELATÓRIO DE DIVULGAÇÕES<br/>RELACIONADAS AO CLIMA", s_title))
        story.append(Paragraph(f"IFRS S1/S2 – Climate-Related Disclosures", s_sub))
        story.append(Spacer(1, 0.5*cm))

        capa_data = [
            ['IDENTIFICAÇÃO', ''],
            ['Entidade', empresa],
            ['CNPJ / Registro', q("q_empresa_cnpj") or "—"],
            ['Setor', q("q_empresa_setor") or "—"],
            ['País', q("q_empresa_pais") or "—"],
            ['Período', periodo],
            ['Consolidação', q("q_consolidacao")],
            ['Responsável', f"{q('q_empresa_responsavel') or '—'} – {q('q_empresa_cargo_responsavel') or ''}"],
        ]
        story.append(_tbl(capa_data, [6*cm, 10*cm]))
        story.append(Spacer(1, 0.8*cm))

        # Sumário executivo
        sum_data = [
            ['SUMÁRIO EXECUTIVO DE EMISSÕES', '', ''],
            ['Métrica', 'Valor', 'Unidade'],
            ['Escopo 1 – Diretas', _fmt_num(ghg['escopo_1_tco2e'], 2), 'tCO₂e'],
            ['Escopo 2 – Energia', _fmt_num(ghg['escopo_2_tco2e'], 2), 'tCO₂e'],
            ['Escopo 3 – Cadeia de Valor', _fmt_num(ghg['escopo_3_tco2e'], 2), 'tCO₂e'],
            ['Total GEE', _fmt_num(ghg['total_tco2e'], 2), 'tCO₂e'],
            ['Intensidade', _fmt_num(ghg['intensidade_media'], 4), 'tCO₂e/t produto'],
            ['Massa Produzida', _fmt_num(ghg['massa_total_t'], 1), 't'],
            ['GEE sob regulação', f"{rd['exposicao_regulatoria']['pct_sob_regulacao']:.1f}%", ''],
        ]
        t = _tbl(sum_data, [6*cm, 5*cm, 5*cm])
        t.setStyle(TableStyle([
            ('SPAN', (0, 0), (-1, 0)),
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, 1), rl_colors.HexColor('#134E4A')),
            ('TEXTCOLOR', (0, 1), (-1, 1), rl_colors.whitesmoke),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 8),
            ('ALIGN', (1, 2), (1, -1), 'RIGHT'),
            ('LINEABOVE', (0, 5), (-1, 5), 1.2, rl_colors.HexColor('#0F766E')),
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
        ]))
        story.append(t)

        story.append(Spacer(1, 0.6*cm))
        story.append(Paragraph(
            f"<i>Gerado pela Calculadora de Pegada de Carbono de Materiais (CMP) em conformidade com "
            f"IFRS S2, GHG Protocol e ISO 14064. Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}.</i>", s_note))
        story.append(PageBreak())

        # ═══ 1. GOVERNANÇA ═══
        story.append(Paragraph("1. GOVERNANÇA (IFRS S2 §5-12)", s_h1))
        if q("q_gov_orgao"):
            gov_data = [['Aspecto', 'Divulgação']]
            gov_data.append(['Órgão supervisor (§6a)', q("q_gov_orgao")])
            gov_data.append(['Frequência de reporte (§6b)', q("q_gov_frequencia")])
            if q("q_gov_comite"):
                gov_data.append(['Comitê dedicado (§6c)', q("q_gov_comite_nome") or "Sim"])
            if q("q_gov_competencias"):
                gov_data.append(['Competências climáticas (§6d)', Paragraph(q("q_gov_competencias"), s_body)])
            if q("q_gov_integracao_estrategia"):
                gov_data.append(['Integração na estratégia (§8)', Paragraph(q("q_gov_integracao_estrategia"), s_body)])
            if q("q_gov_remuneracao_vinculada"):
                gov_data.append(['Remuneração vinculada (§10)', q("q_gov_remuneracao_detalhes") or "Sim"])
            story.append(_tbl(gov_data, [5*cm, 11.5*cm]))
        else:
            story.append(Paragraph(
                "A entidade deve divulgar os órgãos de governança responsáveis pela supervisão "
                "dos riscos climáticos, incluindo competências e processos de decisão.", s_body))
            story.append(Paragraph("<i>Seção a ser complementada com informações da organização.</i>", s_note))
        story.append(Spacer(1, 0.3*cm))

        # ═══ 2. ESTRATÉGIA ═══
        story.append(Paragraph("2. ESTRATÉGIA (IFRS S2 §13-22)", s_h1))

        # Horizontes
        story.append(Paragraph("<b>Definição de Horizontes Temporais (§15)</b>", s_h2))
        hor_data = [['Horizonte', 'Período'],
                    ['Curto prazo', q("q_est_horizonte_curto") or "0-2 anos"],
                    ['Médio prazo', q("q_est_horizonte_medio") or "2-5 anos"],
                    ['Longo prazo', q("q_est_horizonte_longo") or "5-30 anos"]]
        story.append(_tbl(hor_data, [5*cm, 11.5*cm], '#0284C7', '#F0F9FF'))
        story.append(Spacer(1, 0.3*cm))

        story.append(Paragraph("<b>Riscos Climáticos Identificados (§13)</b>", s_h2))
        if riscos:
            r_data = [['Tipo', 'Descrição', 'Exposição', 'Horizonte']]
            for r in riscos:
                r_data.append([
                    Paragraph(r['tipo'], s_body), Paragraph(r['descricao'], s_body),
                    Paragraph(r.get('exposicao', '—'), s_body), r.get('horizonte', '—')
                ])
            story.append(_tbl(r_data, [3.2*cm, 6*cm, 3.8*cm, 3.5*cm], '#DC2626', '#FEF2F2'))
        else:
            story.append(Paragraph("<i>Nenhum risco climático material identificado.</i>", s_note))

        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph("<b>Oportunidades (§13c)</b>", s_h2))
        if oportunidades:
            for o in oportunidades:
                story.append(Paragraph(f"• <b>{o['tipo']}:</b> {o['descricao']}", s_bullet))
        if q("q_est_oportunidades_desc"):
            story.append(Paragraph(f"• <b>Identificada pela organização:</b> {q('q_est_oportunidades_desc')}", s_bullet))

        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph("<b>Impacto Financeiro (§21)</b>", s_h2))
        imp_items = []
        if q("q_est_impacto_receita"):
            imp_items.append(f"Receitas: {q('q_est_impacto_receita')}")
        if q("q_est_impacto_custos"):
            imp_items.append(f"Custos: {q('q_est_impacto_custos')}")
        if q("q_est_impacto_ativos"):
            imp_items.append(f"Ativos: {q('q_est_impacto_ativos')}")
        if imp_items:
            for item in imp_items:
                story.append(Paragraph(f"• {item}", s_bullet))
        else:
            exp = rd['exposicao_regulatoria']
            story.append(Paragraph(_pdf_safe(
                f"A análise identifica {_fmt_num(exp['emissao_cbam_tco2e'], 2)} tCO₂e sob taxação de fronteira e "
                f"{_fmt_num(exp['emissao_local_tco2e'], 2)} tCO₂e sob taxação local ({exp['pct_sob_regulacao']:.1f}%)."), s_body))

        if q("q_est_cenarios"):
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(f"<b>Cenários climáticos (§22):</b> {q('q_est_cenarios_detalhes')}", s_body))

        story.append(PageBreak())

        # ═══ 3. GESTÃO DE RISCOS ═══
        story.append(Paragraph("3. GESTÃO DE RISCOS (IFRS S2 §23-24)", s_h1))
        if q("q_risk_processo"):
            story.append(Paragraph(f"<b>Processo de identificação (§23a):</b> {q('q_risk_processo')}", s_body))
            story.append(Paragraph(f"<b>Frequência:</b> {q('q_risk_frequencia')}", s_body))
            if q("q_risk_integrado"):
                story.append(Paragraph(f"<b>Integração com ERM (§24):</b> {q('q_risk_integrado_desc') or 'Sim'}", s_body))
            if q("q_risk_mitigacao"):
                story.append(Paragraph(f"<b>Ações de mitigação (§24b):</b> {q('q_risk_mitigacao')}", s_body))
        else:
            story.append(Paragraph(
                "Os riscos climáticos são identificados por análise automatizada: intensidade >2× média, "
                "concentração >40%, e exposição regulatória.", s_body))

        story.append(Paragraph("<b>Critérios de Materialidade</b>", s_h2))
        n_units = len(unidades) if unidades else 1
        criteria = [
            "Intensidade de emissão >2× a média do portfólio",
            "Concentração >40% das emissões em uma unidade",
            "Exposição a taxação de carbono (CBAM ou local)",
            f"Emissões absolutas >{_fmt_num(ghg['total_tco2e'] / n_units * 2, 0)} tCO₂e/unidade",
        ]
        for c in criteria:
            story.append(Paragraph(_pdf_safe(f"• {c}"), s_bullet))
        story.append(Paragraph(
            f"<i>Unidades de alto risco identificadas: {len(high_intensity)}</i>", s_note))
        story.append(Spacer(1, 0.3*cm))

        # ═══ 4. MÉTRICAS E METAS ═══
        story.append(Paragraph("4. MÉTRICAS E METAS (IFRS S2 §29-36)", s_h1))
        story.append(Paragraph("<b>Inventário de GEE (§29a)</b>", s_h2))

        ghg_tbl = [
            ['Métrica', 'Valor', 'Unidade', 'Padrão'],
            ['Escopo 1 – Diretas', _fmt_num(ghg['escopo_1_tco2e'], 2), 'tCO₂e', 'GHG Protocol'],
            ['Escopo 2 – Energia (localização)', _fmt_num(ghg['escopo_2_tco2e'], 2), 'tCO₂e', 'GHG Protocol'],
            ['Escopo 3 – Cadeia de Valor', _fmt_num(ghg['escopo_3_tco2e'], 2), 'tCO₂e', 'GHG Protocol'],
            ['Total GEE', _fmt_num(ghg['total_tco2e'], 2), 'tCO₂e', ''],
            ['Intensidade', _fmt_num(ghg['intensidade_media'], 4), 'tCO₂e/t', 'IFRS S2 §29(d)'],
            ['Massa produzida', _fmt_num(ghg['massa_total_t'], 1), 't', ''],
        ]
        t = _tbl(ghg_tbl, [5.5*cm, 3.5*cm, 3*cm, 4.5*cm])
        t.setStyle(TableStyle([
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('LINEABOVE', (0, 4), (-1, 4), 1.2, rl_colors.HexColor('#0F766E')),
            ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

        story.append(Paragraph("<b>Exposição Regulatória (§29b)</b>", s_h2))
        exp = rd['exposicao_regulatoria']
        exp_tbl = [
            ['Métrica', 'Valor', 'Ref.'],
            ['GEE sob regulação de preço', f"{exp['pct_sob_regulacao']:.1f}%", '§29(b)'],
            ['Emissões CBAM/fronteira', f"{_fmt_num(exp['emissao_cbam_tco2e'], 2)} tCO₂e", '§29(b)'],
            ['Emissões taxação local', f"{_fmt_num(exp['emissao_local_tco2e'], 2)} tCO₂e", '§29(b)'],
            ['Unidades alto risco', str(len(high_intensity)), '§29(f)'],
        ]
        story.append(_tbl(exp_tbl, [7*cm, 5.5*cm, 4*cm], '#0284C7', '#F0F9FF'))
        story.append(Spacer(1, 0.3*cm))

        # Metas no PDF
        story.append(Paragraph("<b>Metas Climáticas (§33-36)</b>", s_h2))
        metas = rd.get("metas", {})
        if metas.get("possui"):
            meta_tbl = [
                ['Aspecto', 'Detalhe'],
                ['Tipo de meta', metas.get("tipo", "—")],
                ['Ano-base', str(metas.get("ano_base", "—"))],
                ['Ano-alvo', str(metas.get("ano_alvo", "—"))],
                ['Redução', f"{metas.get('reducao_pct', 0)}%"],
                ['Validação SBTi', 'Sim' if metas.get("sbti") else 'Não'],
                ['Net Zero', metas.get("net_zero", "—") or "—"],
            ]
            story.append(_tbl(meta_tbl, [5*cm, 11.5*cm], '#059669', '#F0FDF4'))
            if q("q_meta_intermediarias"):
                story.append(Paragraph(f"<b>Marcos intermediários:</b> {q('q_meta_intermediarias')}", s_body))
        else:
            story.append(Paragraph("<i>Metas climáticas formais não declaradas.</i>", s_note))

        story.append(PageBreak())

        # ═══ 5. PLANO DE TRANSIÇÃO ═══
        story.append(Paragraph("5. PLANO DE TRANSIÇÃO CLIMÁTICA (IFRS S2 §14)", s_h1))
        if q("q_trans_possui_plano"):
            trans_items = [
                ("Ações previstas", q("q_trans_acoes")),
                ("Investimento", q("q_trans_investimento")),
                ("Tecnologias planejadas", q("q_trans_tecnologias")),
                ("Dependências e premissas", q("q_trans_dependencias")),
            ]
            for label, val in trans_items:
                if val:
                    story.append(Paragraph(f"<b>{label}:</b> {val}", s_body))
        else:
            story.append(Paragraph("<i>A organização não declarou plano de transição climática formal.</i>", s_note))

        tec_set = set()
        for u in unidades:
            if hasattr(u, "Tecnologia") and u.Tecnologia:
                tec_set.add(u.Tecnologia.nome)
        if tec_set:
            story.append(Paragraph(f"<b>Tecnologias atuais em uso:</b> {', '.join(sorted(tec_set))}", s_body))

        story.append(Spacer(1, 0.5*cm))

        # ═══ 6. VERIFICAÇÃO ═══
        story.append(Paragraph("6. VERIFICAÇÃO E ASSEGURAÇÃO", s_h1))
        if q("q_ver_assegurado"):
            ver_tbl = [
                ['Aspecto', 'Detalhe'],
                ['Tipo de asseguração', q("q_ver_tipo")],
                ['Organismo verificador', q("q_ver_auditor") or "—"],
                ['Norma', q("q_ver_norma")],
            ]
            story.append(_tbl(ver_tbl, [5*cm, 11.5*cm], '#7C3AED', '#F5F3FF'))
        else:
            story.append(Paragraph(
                "O inventário de GEE não foi submetido a asseguração por terceiros independentes. "
                "Recomenda-se verificação conforme ISO 14064-3 ou ISAE 3410 para conformidade "
                "completa com IFRS S2.", s_body))

        if q("q_add_offsets"):
            story.append(Paragraph(f"<b>Compensações de carbono:</b> {q('q_add_offsets_desc') or 'Sim'}", s_body))
        if q("q_add_preco_interno"):
            story.append(Paragraph(_pdf_safe(f"<b>Preço interno de carbono:</b> €{q('q_add_preco_valor')}/tCO₂e"), s_body))

        story.append(PageBreak())

        # ═══ ANEXO: UNIDADES ═══
        story.append(Paragraph("ANEXO A: DETALHAMENTO POR UNIDADE PRODUTIVA", s_h1))
        sorted_units = sorted(unidades, key=lambda u: u.IntensidadeEmissao * u.MassaOutput, reverse=True)[:15]
        u_data = [['ID', 'Nome', 'Local', 'E1', 'E2', 'E3', 'Total', 'Tax.']]
        for u in sorted_units:
            tax = []
            if u.TaxacaoFronteira: tax.append("CBAM")
            if u.TaxacaoLocal: tax.append("Local")
            u_data.append([
                u.ID_ELO, Paragraph(u.Nome[:25], s_body), u.Localizacao[:15] if u.Localizacao else "—",
                _fmt_num(u.IntensidadeEmissaoEscopo1 * u.MassaOutput, 1),
                _fmt_num(u.IntensidadeEmissaoEscopo2 * u.MassaOutput, 1),
                _fmt_num(u.IntensidadeEmissaoEscopo3 * u.MassaOutput, 1),
                _fmt_num(u.IntensidadeEmissao * u.MassaOutput, 1),
                ", ".join(tax) or "—",
            ])
        story.append(_tbl(u_data, [1.8*cm, 3.2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 1.5*cm], '#334155'))
        story.append(Paragraph(_pdf_safe(
            f"<i>Listagem das {len(sorted_units)} maiores unidades por emissão total. "
            f"Valores em tCO₂e. Total de unidades no modelo: {len(unidades)}.</i>"), s_note))

        # ═══ DISCLAIMER ═══
        story.append(Spacer(1, 1*cm))
        story.append(HRFlowable(width="100%", color=rl_colors.HexColor('#E2E8F0')))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            "<b>Disclaimer:</b> Este relatório foi gerado automaticamente com base nos dados inseridos na "
            "Calculadora de Pegada de Carbono de Materiais e nas respostas ao questionário IFRS. "
            "A conformidade total com IFRS S1/S2 requer revisão por profissionais especializados em "
            "finanças sustentáveis e auditoria de GEE. Os dados devem ser verificados por terceiros "
            "independentes antes da publicação oficial.", s_note))
        if q("q_add_notas"):
            story.append(Paragraph(f"<b>Observações adicionais:</b> {q('q_add_notas')}", s_note))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            f"<i>Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
            f"Padrões: IFRS S1, IFRS S2, GHG Protocol, ISO 14064 | v2.1</i>", s_note))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    # ════════════════════════════════════════════════════════════════
    #  COMPARATIVO MULTI-ANO
    # ════════════════════════════════════════════════════════════════
    def _render_comparativo(self):
        """Renderiza aba de análise comparativa entre anos."""
        from core.context import AppContext
        from core.calc.comparativo import (
            pivot_intensidade,
            calcular_deltas,
            calcular_variacao_pct,
            resumo_comparativo,
            dados_grafico_barras_comparativo,
            dados_grafico_linha_evolucao,
        )
        from core.periodos import format_periodo

        ctx = AppContext.get()
        unidades = st.session_state.unidades

        st.markdown("### 🔄 Análise Comparativa Multi-Ano")

        if not ctx.modo_comparacao or len(ctx.anos_selecionados) < 2:
            st.info(
                "Ative o **Modo comparativo** na barra lateral e selecione 2+ anos "
                "para visualizar análises comparativas.",
                icon="ℹ️",
            )
            st.markdown(
                "**Como usar:**\n"
                "1. Na barra lateral, ative o toggle \"Modo comparativo\"\n"
                "2. Selecione múltiplos anos ou use uma expressão como `2020-2025`\n"
                "3. Os gráficos e tabelas comparativas aparecerão aqui"
            )
            return

        anos = ctx.anos_selecionados
        st.caption(f"Comparando: **{format_periodo(anos)}**")

        # ── Unidades de exibição ──
        _mu = _norm_unit(st.session_state.get("mass_unit", "t"))
        co2e_lbl = co2e_label(_mu)
        int_lbl = co2e_intensity_label(_mu)
        c2e = lambda v: convert_co2e(v, _mu)
        c2i = lambda v: convert_co2e(v, "t")

        # --- Resumo geral ---
        resumo = resumo_comparativo(unidades, anos)
        if not resumo.get("dados"):
            st.warning("Nenhum dado encontrado para os anos selecionados.")
            return

        st.markdown("#### 📊 Resumo por Ano")
        cols = st.columns(len(anos))
        for i, ano in enumerate(sorted(anos)):
            dados_ano = resumo["dados"].get(ano, {})
            with cols[i]:
                st.metric(f"Ano {ano}", f"{c2e(dados_ano.get('emissao_total', 0)):,.4f} {co2e_lbl}")
                st.caption(
                    f"Unidades: {dados_ano.get('total_unidades', 0)} · "
                    f"Massa: {dados_ano.get('massa_total', 0):,.1f} t"
                )

        # Deltas entre anos consecutivos
        if resumo.get("deltas"):
            st.markdown("#### ↕️ Variação entre Períodos")
            for d in resumo["deltas"]:
                delta_val = d["delta_emissao"]
                pct_val = d["variacao_pct"]
                arrow = "🔺" if delta_val > 0 else "🔻" if delta_val < 0 else "➖"
                st.markdown(
                    f"**{d['de']} → {d['para']}**: {arrow} "
                    f"Δ = {c2e(delta_val):+,.4f} {co2e_lbl} ({pct_val:+.2f}%)"
                )

        st.markdown("---")

        # --- Seletor de métrica ---
        metrica = st.selectbox(
            "Métrica para comparação:",
            options=[
                "IntensidadeEmissao",
                "IntensidadeEscopo1",
                "IntensidadeEscopo2",
                "IntensidadeEscopo3",
                "Pegada",
            ],
            format_func=lambda x: {
                "IntensidadeEmissao": f"Intensidade Total ({int_lbl})",
                "IntensidadeEscopo1": "Intensidade Escopo 1",
                "IntensidadeEscopo2": "Intensidade Escopo 2",
                "IntensidadeEscopo3": "Intensidade Escopo 3",
                "Pegada": "Pegada Acumulada",
            }.get(x, x),
            key="_comp_metrica",
        )

        tab_chart, tab_table = st.tabs(["📊 Gráficos", "📋 Tabelas"])

        with tab_chart:
            # Gráfico de barras agrupadas
            st.markdown("##### Barras Comparativas por Unidade")
            dados_barras = dados_grafico_barras_comparativo(unidades, anos, metrica)
            if dados_barras["labels"]:
                fig_bar = go.Figure()
                colors = px.colors.qualitative.Set2
                for i, (ano_label, valores) in enumerate(dados_barras["series"].items()):
                    fig_bar.add_trace(go.Bar(
                        name=str(ano_label),
                        x=dados_barras["labels"],
                        y=valores,
                        marker_color=colors[i % len(colors)],
                        text=[f"{v:.4f}" for v in valores],
                        textposition="outside",
                    ))
                fig_bar.update_layout(
                    barmode="group",
                    yaxis_title=int_lbl,
                    height=400,
                    margin=dict(l=20, r=20, t=30, b=40),
                    plot_bgcolor="white",
                    legend=dict(orientation="h", y=1.05, x=0.5, xanchor="center"),
                )
                st.plotly_chart(fig_bar, use_container_width=True)
            else:
                st.info("Sem dados para gráfico de barras.")

            # Gráfico de linha (evolução)
            st.markdown("##### Evolução Temporal")
            dados_linha = dados_grafico_linha_evolucao(unidades, anos, metrica)
            if dados_linha.get("anos"):
                fig_line = go.Figure()
                colors_l = px.colors.qualitative.Plotly
                for i, (nome, valores) in enumerate(dados_linha["series"].items()):
                    fig_line.add_trace(go.Scatter(
                        x=dados_linha["anos"],
                        y=valores,
                        mode="lines+markers",
                        name=nome,
                        line=dict(color=colors_l[i % len(colors_l)], width=2),
                        marker=dict(size=8),
                    ))
                fig_line.update_layout(
                    xaxis_title="Ano",
                    yaxis_title=int_lbl,
                    height=400,
                    margin=dict(l=20, r=20, t=30, b=40),
                    plot_bgcolor="white",
                    legend=dict(orientation="h", y=1.05, x=0.5, xanchor="center"),
                )
                st.plotly_chart(fig_line, use_container_width=True)
            else:
                st.info("Sem dados para gráfico de evolução.")

            # Gráfico de barras por escopo × ano
            st.markdown("##### Emissões por Escopo × Ano")
            dados_esc = dados_grafico_barras_comparativo(unidades, anos, por="escopo")
            if dados_esc["labels"]:
                fig_esc = go.Figure()
                esc_colors = {"Escopo 1": _COLORS["scope1"], "Escopo 2": _COLORS["scope2"], "Escopo 3": _COLORS["scope3"]}
                for escopo, valores in dados_esc["series"].items():
                    fig_esc.add_trace(go.Bar(
                        name=escopo,
                        x=dados_esc["labels"],
                        y=valores,
                        marker_color=esc_colors.get(escopo, "#888"),
                        text=[f"{v:.4f}" for v in valores],
                        textposition="outside",
                    ))
                fig_esc.update_layout(
                    barmode="group",
                    yaxis_title=co2e_lbl,
                    height=400,
                    margin=dict(l=20, r=20, t=30, b=40),
                    plot_bgcolor="white",
                    legend=dict(orientation="h", y=1.05, x=0.5, xanchor="center"),
                )
                st.plotly_chart(fig_esc, use_container_width=True)

        with tab_table:
            # Pivot table
            st.markdown("##### Pivot: Unidade × Ano")
            pivot_df = pivot_intensidade(unidades, anos, metrica)
            if not pivot_df.empty:
                st.dataframe(pivot_df.style.format("{:.4f}"), use_container_width=True)
            else:
                st.info("Sem dados para pivot table.")

            # Deltas
            st.markdown("##### Variação Absoluta (Δ)")
            deltas_df = calcular_deltas(unidades, anos, metrica)
            if not deltas_df.empty:
                st.dataframe(
                    deltas_df.style.format("{:.4f}").applymap(
                        lambda v: "color: red" if isinstance(v, (int, float)) and v > 0 else "color: green",
                    ),
                    use_container_width=True,
                )
            else:
                st.info("Necessário 2+ anos para calcular deltas.")

            # Variação %
            st.markdown("##### Variação Percentual (%)")
            pct_df = calcular_variacao_pct(unidades, anos, metrica)
            if not pct_df.empty:
                st.dataframe(
                    pct_df.style.format("{:.2f}%").applymap(
                        lambda v: "color: red" if isinstance(v, (int, float)) and v > 0 else "color: green",
                    ),
                    use_container_width=True,
                )

        # ── Download bar ──
        comp_data = {
            "anos": list(anos),
            "resumo": resumo.get("dados", {}),
            "deltas": resumo.get("deltas", []),
        }
        # Build pivot MD representation
        if not pivot_df.empty:
            lines = []
            cols = ["Unidade"] + [str(c) for c in pivot_df.columns]
            lines.append("| " + " | ".join(cols) + " |")
            lines.append("| " + " | ".join([":-"] + ["-:"] * (len(cols) - 1)) + " |")
            for idx, row in pivot_df.iterrows():
                vals = [str(idx)] + [f"{v:.4f}" for v in row]
                lines.append("| " + " | ".join(vals) + " |")
            comp_data["pivot_md"] = "\n".join(lines) + "\n"
        else:
            comp_data["pivot_md"] = ""

        cfg = get_report_config()
        md_comp = generate_md_comparativo(comp_data, cfg)
        pdf_comp = generate_pdf_comparativo(comp_data, cfg)
        render_download_bar("comparativo", md_comp, pdf_comp, filename_base="Comparativo_MultiAno")

    # ════════════════════════════════════════════════════════════════
    #  PAINEL GERAL (mantido)
    # ════════════════════════════════════════════════════════════════
    def _render_painel_geral(self):
        unidades = st.session_state.unidades
        edges = st.session_state.edges
        total_unidades = len(unidades)
        emissao_total = sum(u.IntensidadeEmissao * u.MassaOutput for u in unidades)
        massa_total_output = sum(u.MassaOutput for u in unidades)
        intensidade_media = emissao_total / massa_total_output if massa_total_output > 0 else 0
        esc1 = sum(u.IntensidadeEmissaoEscopo1 * u.MassaOutput for u in unidades)
        esc2 = sum(u.IntensidadeEmissaoEscopo2 * u.MassaOutput for u in unidades)
        esc3 = sum(u.IntensidadeEmissaoEscopo3 * u.MassaOutput for u in unidades)
        unidades_taxadas = sum(1 for u in unidades if u.TaxacaoFronteira or u.TaxacaoLocal)
        pct_taxadas = (unidades_taxadas / total_unidades * 100) if total_unidades else 0

        # ── Unit helpers (values are in kgCO₂e internally) ──
        _mu = get_default_mass_unit_from_session(st.session_state)
        co2e_lbl = co2e_label(_mu)
        int_lbl = co2e_intensity_label(_mu)
        c2e = lambda v: convert_co2e(v, _mu)
        c2i = lambda v: convert_co2e(v, "t")  # intensity always ÷1000

        st.markdown("### Indicadores-Chave")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Emissão Total", f"{c2e(emissao_total):,.2f} {co2e_lbl}")
        c2.metric("Intensidade Média", f"{c2i(intensidade_media):,.4f} {int_lbl}")
        c3.metric("Massa Produzida", f"{massa_total_output:,.1f} t")
        c4.metric("Unidades", total_unidades)
        c5.metric("Taxadas", f"{unidades_taxadas} ({pct_taxadas:.0f}%)")
        st.markdown("---")

        st.markdown("### Emissões por Escopo (GHG Protocol)")
        col_chart, col_table = st.columns([2, 1])
        with col_chart:
            fig = go.Figure(go.Bar(
                x=["Escopo 1", "Escopo 2", "Escopo 3"], y=[c2e(esc1), c2e(esc2), c2e(esc3)],
                marker_color=[_COLORS["scope1"], _COLORS["scope2"], _COLORS["scope3"]],
                text=[f"{c2e(v):,.2f}" for v in [esc1, esc2, esc3]], textposition="outside",
            ))
            fig.update_layout(yaxis_title=co2e_lbl, height=350, margin=dict(l=20,r=20,t=30,b=20), plot_bgcolor="white")
            st.plotly_chart(fig, use_container_width=True)
        with col_table:
            total_e = esc1 + esc2 + esc3
            pct = lambda v: f"{v/total_e*100:.1f}%" if total_e > 0 else "0%"
            st.markdown(f"""
| Escopo | {co2e_lbl} | % |
|--------|------:|:-:|
| 🔴 Escopo 1 | {c2e(esc1):,.2f} | {pct(esc1)} |
| 🟡 Escopo 2 | {c2e(esc2):,.2f} | {pct(esc2)} |
| 🔵 Escopo 3 | {c2e(esc3):,.2f} | {pct(esc3)} |
| **Total** | **{c2e(total_e):,.2f}** | **100%** |
""")

        st.markdown("---")
        col_top, col_dist = st.columns(2)
        with col_top:
            st.markdown("#### 🏭 Top 5 Emissores")
            top5 = sorted(unidades, key=lambda x: x.IntensidadeEmissao * x.MassaOutput, reverse=True)[:5]
            fig_t = go.Figure(go.Bar(
                y=[u.ID_ELO for u in top5], x=[c2e(u.IntensidadeEmissao*u.MassaOutput) for u in top5],
                orientation="h", marker_color=_COLORS["primary"],
                text=[f"{c2e(u.IntensidadeEmissao*u.MassaOutput):,.2f}" for u in top5], textposition="outside",
            ))
            fig_t.update_layout(xaxis_title=co2e_lbl, height=300, margin=dict(l=60,r=40,t=10,b=20),
                               yaxis=dict(autorange="reversed"), plot_bgcolor="white")
            st.plotly_chart(fig_t, use_container_width=True)
        with col_dist:
            st.markdown("#### 📍 Por Localização")
            loc_d = {}
            for u in unidades:
                l = u.Localizacao or "Sem local"
                loc_d[l] = loc_d.get(l, 0) + c2e(u.IntensidadeEmissao * u.MassaOutput)
            if loc_d:
                fig_p = go.Figure(go.Pie(labels=list(loc_d.keys()), values=list(loc_d.values()), hole=0.4,
                                         textinfo="label+percent", marker_colors=px.colors.qualitative.Set2))
                fig_p.update_layout(height=300, margin=dict(l=10,r=10,t=10,b=10), showlegend=False)
                st.plotly_chart(fig_p, use_container_width=True)

        st.markdown("---")
        st.markdown("### Emissões por Escopo por Unidade")
        sorted_all = sorted(unidades, key=lambda x: x.IntensidadeEmissao*x.MassaOutput, reverse=True)
        fig_s = go.Figure()
        for lbl, clr, acc in [
            ("Escopo 1", _COLORS["scope1"], lambda u: c2e(u.IntensidadeEmissaoEscopo1*u.MassaOutput)),
            ("Escopo 2", _COLORS["scope2"], lambda u: c2e(u.IntensidadeEmissaoEscopo2*u.MassaOutput)),
            ("Escopo 3", _COLORS["scope3"], lambda u: c2e(u.IntensidadeEmissaoEscopo3*u.MassaOutput)),
        ]:
            fig_s.add_trace(go.Bar(x=[u.ID_ELO for u in sorted_all], y=[acc(u) for u in sorted_all],
                                   name=lbl, marker_color=clr))
        fig_s.update_layout(barmode="stack", yaxis_title=co2e_lbl, height=400,
                            margin=dict(l=20,r=20,t=30,b=40), plot_bgcolor="white",
                            legend=dict(orientation="h", y=1.02, x=0.5, xanchor="center"))
        st.plotly_chart(fig_s, use_container_width=True)

        # ── Download bar ──
        painel_data = {
            "emissao_total": emissao_total,
            "intensidade_media": intensidade_media,
            "massa_total": massa_total_output,
            "total_unidades": total_unidades,
            "unidades_taxadas": unidades_taxadas,
            "pct_taxadas": pct_taxadas,
            "esc1": esc1, "esc2": esc2, "esc3": esc3,
            "top5": [{"id": u.ID_ELO, "emissao": u.IntensidadeEmissao * u.MassaOutput} for u in top5],
            "por_localizacao": loc_d,
            "todas_unidades": [
                {
                    "id": u.ID_ELO,
                    "e1": u.IntensidadeEmissaoEscopo1 * u.MassaOutput,
                    "e2": u.IntensidadeEmissaoEscopo2 * u.MassaOutput,
                    "e3": u.IntensidadeEmissaoEscopo3 * u.MassaOutput,
                    "total": u.IntensidadeEmissao * u.MassaOutput,
                }
                for u in sorted_all
            ],
        }
        cfg = get_report_config()
        md_painel = generate_md_painel_geral(painel_data, cfg)
        pdf_painel = generate_pdf_painel_geral(painel_data, cfg)
        render_download_bar("painel", md_painel, pdf_painel, filename_base="Painel_Geral_Emissoes")

    # ════════════════════════════════════════════════════════════════
    #  DIAGRAMA SANKEY (mantido)
    # ════════════════════════════════════════════════════════════════
    def _render_sankey_diagram(self):
        st.markdown("### Diagrama Sankey – Fluxo de Emissões")
        _mu_sk = get_default_mass_unit_from_session(st.session_state)
        _co2e_lbl_sk = co2e_label(_mu_sk)
        _c2e_sk = lambda v: convert_co2e(v, _mu_sk)
        dims = self._get_dimensoes_disponiveis()
        with st.expander("⚙️ Configurações do Sankey", expanded=True):
            c1, c2 = st.columns([3, 1])
            with c1:
                flux = st.multiselect("Dimensões:", list(dims.keys()),
                                       default=list(dims.keys())[:2] if len(dims) >= 2 else list(dims.keys())[:1],
                                       key="sankey_flux_ordem")
            with c2:
                show_u = st.checkbox("Mostrar unidades", False, key="sankey_mostrar_unidades")
                alt = st.slider("Altura (px)", 400, 1200, 700, 50, key="sankey_altura")
        if len(flux) < 1:
            st.info("Selecione ao menos 1 dimensão.", icon="ℹ️"); return
        labels, src, tgt, val, clrs, hover = self._build_aggregated_sankey(
            flux, "Intensidade de Emissão (tCO2e)", "Auto", st.session_state.unidades, show_u)
        if not val:
            st.warning("Sem fluxos válidos."); return
        nc = self._get_node_colors_advanced(labels, "Auto", flux)
        fig = go.Figure(go.Sankey(
            node=dict(pad=20, thickness=25, label=labels, color=nc, line=dict(width=0.5, color="white")),
            link=dict(source=src, target=tgt, value=val, color=clrs, customdata=hover,
                      hovertemplate="%{customdata}<extra></extra>"),
            textfont=dict(color="black", size=12, family="Arial"),
        ))
        fig.update_layout(title_text=f"Fluxo: {' → '.join(flux)}" + (" → Unidades" if show_u else ""),
                          font=dict(size=14, color="black"), height=alt, margin=dict(l=20,r=120,t=60,b=40))
        st.plotly_chart(fig, use_container_width=True)
        c1, c2, c3 = st.columns(3)
        c1.metric("Nós", len(labels)); c2.metric("Arcos", len(val))
        c3.metric("Emissão Total", f"{_c2e_sk(sum(val)):,.2f} {_co2e_lbl_sk}")

    # ════════════════════════════════════════════════════════════════
    #  MATRIZ DE RISCO (mantida)
    # ════════════════════════════════════════════════════════════════
    def _render_risk_matrix(self, unidades, intensidade_media):
        if not unidades: return
        _mu_rm = get_default_mass_unit_from_session(st.session_state)
        _co2e_lbl_rm = co2e_label(_mu_rm)
        _int_lbl_rm = co2e_intensity_label(_mu_rm)
        _c2e_rm = lambda v: convert_co2e(v, _mu_rm)
        _c2i_rm = lambda v: convert_co2e(v, "t")
        x, y, t, sz, cl = [], [], [], [], []
        for u in unidades:
            em = u.IntensidadeEmissao * u.MassaOutput
            i = u.IntensidadeEmissao
            x.append(_c2e_rm(em)); y.append(_c2i_rm(i)); t.append(u.ID_ELO)
            sz.append(max(10, min(50, u.MassaOutput / max(1, max(un.MassaOutput for un in unidades)) * 40 + 10)))
            tx = u.TaxacaoFronteira or u.TaxacaoLocal
            if i > intensidade_media * 2 and tx: cl.append(_COLORS["danger"])
            elif i > intensidade_media * 2 or tx: cl.append(_COLORS["warning"])
            elif i > intensidade_media: cl.append(_COLORS["info"])
            else: cl.append(_COLORS["success"])
        fig = go.Figure(go.Scatter(x=x, y=y, mode="markers+text", text=t, textposition="top center",
                                    textfont=dict(size=9), marker=dict(size=sz, color=cl, opacity=0.8,
                                    line=dict(width=1, color="white")),
                                    hovertemplate=f"%{{text}}<br>%{{x:,.2f}} {_co2e_lbl_rm}<br>%{{y:,.4f}} {_int_lbl_rm}<extra></extra>"))
        fig.add_hline(y=_c2i_rm(intensidade_media), line_dash="dash", line_color=_COLORS["muted"],
                      annotation_text=f"Média: {_c2i_rm(intensidade_media):.4f}")
        fig.add_hline(y=_c2i_rm(intensidade_media)*2, line_dash="dot", line_color=_COLORS["danger"],
                      annotation_text="Limiar 2×")
        fig.update_layout(xaxis_title=f"Emissão Total ({_co2e_lbl_rm})", yaxis_title=f"Intensidade ({_int_lbl_rm})",
                          height=380, margin=dict(l=20,r=20,t=30,b=20), plot_bgcolor="white")
        st.plotly_chart(fig, use_container_width=True)
        st.markdown("🔴 Alto · 🟡 Médio · 🔵 Atenção · 🟢 Baixo — *tamanho ∝ massa*")

    # ════════════════════════════════════════════════════════════════
    #  IFRS INFO (sem dados)
    # ════════════════════════════════════════════════════════════════
    def _render_ifrs_info_only(self):
        st.markdown("---")
        with st.expander("ℹ️ Sobre os Padrões IFRS S1/S2", expanded=True):
            st.markdown("""
### IFRS S1 – Requisitos Gerais de Sustentabilidade
Estabelece requisitos gerais de divulgação: Governança, Estratégia, Gestão de Riscos, Métricas e Metas.

### IFRS S2 – Divulgações Climáticas
| Pilar | Requisitos | Ref. |
|-------|-----------|------|
| 🏛️ Governança | Supervisão de riscos climáticos | §5-12 |
| 📐 Estratégia | Riscos físicos/transição, cenários | §13-22 |
| 🛡️ Gestão de Riscos | Identificação e avaliação | §23-24 |
| 📊 Métricas | GEE por escopo, metas | §29-36 |

### GHG Protocol
🔴 **Escopo 1** – Diretas · 🟡 **Escopo 2** – Energia · 🔵 **Escopo 3** – Cadeia de valor
""")

    # ════════════════════════════════════════════════════════════════
    #  ANÁLISE POR UNIDADE (mantida)
    # ════════════════════════════════════════════════════════════════
    def _render_analise_por_unidade(self):
        st.markdown("### Análise Detalhada por Unidade")
        unidades = st.session_state.unidades
        edges = st.session_state.edges
        _mu_au = get_default_mass_unit_from_session(st.session_state)
        _co2e_lbl_au = co2e_label(_mu_au)
        _int_lbl_au = co2e_intensity_label(_mu_au)
        _c2e_au = lambda v: convert_co2e(v, _mu_au)
        _c2i_au = lambda v: convert_co2e(v, "t")
        dados = []
        for u in unidades:
            ent = sum(1 for e in edges if e["target"] == u.ID_ELO)
            sai = sum(1 for e in edges if e["source"] == u.ID_ELO)
            tx = []
            if u.TaxacaoFronteira: tx.append("Fronteira")
            if u.TaxacaoLocal: tx.append("Local")
            dados.append({"ID": u.ID_ELO, "Nome": u.Nome, "Local": u.Localizacao,
                          "Massa In (t)": round(u.MassaInput, 2), "Massa Out (t)": round(u.MassaOutput, 2),
                          f"E1 ({_co2e_lbl_au})": round(_c2e_au(u.IntensidadeEmissaoEscopo1*u.MassaOutput), 2),
                          f"E2 ({_co2e_lbl_au})": round(_c2e_au(u.IntensidadeEmissaoEscopo2*u.MassaOutput), 2),
                          f"E3 ({_co2e_lbl_au})": round(_c2e_au(u.IntensidadeEmissaoEscopo3*u.MassaOutput), 2),
                          f"Total ({_co2e_lbl_au})": round(_c2e_au(u.IntensidadeEmissao*u.MassaOutput), 2),
                          "Intensidade": round(_c2i_au(u.IntensidadeEmissao), 4),
                          "Pegada": round(_c2i_au(u.Pegada), 4),
                          "Taxação": ", ".join(tx) or "—", "In": ent, "Out": sai})
        _total_col = f"Total ({_co2e_lbl_au})"
        df = pd.DataFrame(dados)
        c1, c2, c3 = st.columns(3)
        with c1: ft = st.multiselect("Taxação:", ["Fronteira","Local","—"], key="ft_v3")
        with c2: ob = st.selectbox("Ordenar:", [_total_col,"Intensidade","Pegada","ID"], key="ob_v3")
        with c3: asc = st.checkbox("Crescente", False, key="asc_v3")
        if ft: df = df[df["Taxação"].apply(lambda x: any(f in x for f in ft))]
        if ob != "ID": df = df.sort_values(ob, ascending=asc)
        st.dataframe(df, use_container_width=True, hide_index=True, height=450)
        st.markdown("---")
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Unidades", len(df)); c2.metric("Emissão", f"{df[_total_col].sum():,.2f} {_co2e_lbl_au}")
        c3.metric("Intens. Média", f"{df['Intensidade'].mean():,.4f} {_int_lbl_au}"); c4.metric("Pegada Média", f"{df['Pegada'].mean():,.4f}")

        # ── Download bar ──
        analise_data = {
            "unidades": df.to_dict("records"),
            "total": len(df),
            "emissao_total": df[_total_col].sum(),
            "intensidade_media": df["Intensidade"].mean(),
            "pegada_media": df["Pegada"].mean(),
        }
        cfg = get_report_config()
        md_analise = generate_md_analise_unidade(analise_data, cfg)
        pdf_analise = generate_pdf_analise_unidade(analise_data, cfg)
        render_download_bar("analise_unidade", md_analise, pdf_analise, filename_base="Analise_por_Unidade")

    # ════════════════════════════════════════════════════════════════
    #  FUNÇÕES AUXILIARES SANKEY
    # ════════════════════════════════════════════════════════════════
    def _get_dimensoes_disponiveis(self):
        dims = {}
        locs, nomes, tecs, cons, escs, faixas = set(), set(), set(), set(), set(), set()
        for u in st.session_state.unidades:
            if u.Localizacao: locs.add(u.Localizacao)
            if u.Nome: nomes.add(u.Nome)
            if hasattr(u,"Tecnologia") and u.Tecnologia: tecs.add(u.Tecnologia.nome)
            if hasattr(u,"Consumiveis") and u.Consumiveis:
                for c in u.Consumiveis:
                    if isinstance(c, dict):
                        if c.get("nome"): cons.add(c["nome"])
                        if c.get("escopo"): escs.add(f"Escopo {c['escopo']}")
                        f = c.get("fator", 0)
                        if f > 0:
                            if f < 0.5: faixas.add("Baixa")
                            elif f < 2: faixas.add("Média")
                            elif f < 5: faixas.add("Alta")
                            else: faixas.add("Muito Alta")
            elif hasattr(u, "Inputs") and u.Inputs:
                for item in u.Inputs:
                    if not isinstance(item, dict):
                        continue
                    produto = str(item.get("produto_id", "") or "").strip()
                    if produto:
                        cons.add(produto)
        if len(locs) > 1: dims["Localização"] = locs
        if len(tecs) > 1: dims["Tecnologia"] = tecs
        if len(nomes) > 1: dims["Nome"] = nomes
        if len(cons) > 1: dims["Tipo de Consumível"] = cons
        if len(escs) > 1: dims["Escopo do Consumível"] = escs
        if len(faixas) > 1: dims["Intensidade do Consumível"] = faixas
        return dims

    def _get_dimensao_valor(self, u, d):
        if d == "Localização": return getattr(u, "Localizacao", None)
        if d == "Tecnologia": return u.Tecnologia.nome if hasattr(u,"Tecnologia") and u.Tecnologia else None
        if d == "Nome": return getattr(u, "Nome", None)
        if d == "Unidade": return getattr(u, "ID_ELO", None)
        return None

    def _get_dimensao_valor_consumivel(self, u, c, d, i):
        if d == "Tipo de Consumível": return c.get("nome", f"C{i+1}")
        if d == "Escopo do Consumível":
            e = c.get("escopo", ""); return f"Escopo {e}" if e else "Sem Escopo"
        if d == "Intensidade do Consumível":
            f = c.get("fator", 0)
            if f < 0.5: return "Baixa"
            if f < 2: return "Média"
            if f < 5: return "Alta"
            return "Muito Alta"
        return None

    def _get_valor_fluxo(self, u, v):
        if "Pegada" in v: return u.Pegada
        if "Massa" in v: return u.MassaOutput
        if "Intensidade" in v: return u.IntensidadeEmissao
        if "Emissão" in v: return u.IntensidadeEmissao * u.MassaOutput
        return 0

    def _build_aggregated_sankey(self, flux, val_exib, cor, units, show_u):
        dc = ["Tipo de Consumível", "Escopo do Consumível", "Intensidade do Consumível"]
        _mu_san = get_default_mass_unit_from_session(st.session_state)
        _co2e_lbl_san = co2e_label(_mu_san)
        _c2e_san = lambda v: convert_co2e(v, _mu_san)
        has_dc = any(d in flux for d in dc)
        rows = []
        for u in units:
            if has_dc and hasattr(u,"Consumiveis") and u.Consumiveis:
                for i, c in enumerate(u.Consumiveis):
                    if isinstance(c, dict):
                        r = {"ID_ELO": u.ID_ELO, "u": u}
                        fc = u.ConsumoEspecifico[i] if hasattr(u,"ConsumoEspecifico") and i < len(u.ConsumoEspecifico) else 0
                        fe = c.get("fator", 0)
                        for d in flux:
                            r[d] = self._get_dimensao_valor_consumivel(u,c,d,i) if d in dc else (self._get_dimensao_valor(u,d) or f"Sem {d}")
                        r["valor"] = fc * fe * u.MassaOutput; rows.append(r)
            elif has_dc and hasattr(u, "Inputs") and u.Inputs:
                try:
                    from core.calc.fatores import FatorIndex
                    fator_idx = FatorIndex(st.session_state.get("fatores_emissao", []))
                except Exception:
                    fator_idx = None

                try:
                    ano_ref = int(float(str(getattr(u, "Periodo", "")).strip()))
                except (TypeError, ValueError):
                    ano_ref = None

                massa_output = float(getattr(u, "MassaOutput", 0.0) or 0.0)
                for i, item in enumerate(u.Inputs):
                    if not isinstance(item, dict):
                        continue

                    nome = str(item.get("produto_id", "") or "").strip()
                    if not nome:
                        continue

                    try:
                        quantidade = float(item.get("quantidade", 0.0) or 0.0)
                    except (TypeError, ValueError):
                        quantidade = 0.0
                    consumo_espec = (quantidade / massa_output) if massa_output > 0 else 0.0

                    fator = 0.0
                    escopo = "SCOPE 1"
                    if fator_idx:
                        for esc_try in ("1", "2", "3"):
                            d = fator_idx.get_fator_dict(nome, esc_try, ano=ano_ref)
                            if d is None:
                                d = fator_idx.get_fator_dict(nome, esc_try, ano=None)
                            if d is not None:
                                fator = float(d.get("fator_emissao", 0.0))
                                escopo = str(d.get("escopo", f"SCOPE {esc_try}"))
                                break

                    c = {"nome": nome, "escopo": escopo, "fator": fator}
                    r = {"ID_ELO": u.ID_ELO, "u": u}
                    for d in flux:
                        r[d] = self._get_dimensao_valor_consumivel(u, c, d, i) if d in dc else (self._get_dimensao_valor(u, d) or f"Sem {d}")
                    r["valor"] = consumo_espec * fator * massa_output
                    rows.append(r)
            else:
                r = {"ID_ELO": u.ID_ELO, "u": u}
                for d in flux: r[d] = self._get_dimensao_valor(u,d) or f"Sem {d}"
                r["valor"] = self._get_valor_fluxo(u, val_exib); rows.append(r)
        if not rows: return [],[],[],[],[],[]
        df = pd.DataFrame(rows); nodes, nm = [], {}
        for col in flux:
            for v in df[col].dropna().unique():
                lbl = f"{col}: {v}"
                if lbl not in nm: nm[lbl] = len(nodes); nodes.append(lbl)
        if show_u:
            for eid in df["ID_ELO"].unique():
                if eid not in nm: nm[eid] = len(nodes); nodes.append(eid)
        src, tgt, val, hov = [], [], [], []
        for i in range(len(flux)-1):
            a, b = flux[i], flux[i+1]
            g = df.groupby([a,b])["valor"].sum().reset_index()
            for _,r in g.iterrows():
                if r["valor"] > 0:
                    ol, dl = f"{a}: {r[a]}", f"{b}: {r[b]}"
                    if ol in nm and dl in nm:
                        src.append(nm[ol]); tgt.append(nm[dl]); val.append(r["valor"])
                        hov.append(f"{ol}<br>→ {dl}<br>{_c2e_san(r['valor']):,.2f} {_co2e_lbl_san}")
        if show_u:
            ud = flux[-1]; gf = df.groupby([ud,"ID_ELO"])["valor"].sum().reset_index()
            for _,r in gf.iterrows():
                if r["valor"] > 0:
                    ol, dl = f"{ud}: {r[ud]}", r["ID_ELO"]
                    if ol in nm and dl in nm:
                        src.append(nm[ol]); tgt.append(nm[dl]); val.append(r["valor"])
                        hov.append(f"{ol}<br>→ {dl}<br>{_c2e_san(r['valor']):,.2f} {_co2e_lbl_san}")

        total_val = sum(val)
        if total_val > 0:
            hov = [f"{h}<br>Contribuição: {(v / total_val) * 100:.2f}%" for h, v in zip(hov, val)]

        return nodes, src, tgt, val, ["rgba(150,150,150,0.3)"]*len(val), hov

    def _get_node_colors_advanced(self, labels, cor, flux):
        pal = px.colors.qualitative.Plotly + px.colors.qualitative.Set2
        dm = {d: pal[i % len(pal)] for i, d in enumerate(flux)}
        out = []
        for lbl in labels:
            if ":" in lbl:
                out.append(dm.get(lbl.split(":")[0], "rgba(100,100,100,0.8)"))
            else:
                u = next((u for u in st.session_state.unidades if u.ID_ELO == lbl), None)
                if u:
                    mx = max((un.IntensidadeEmissao for un in st.session_state.unidades), default=1)
                    n = u.IntensidadeEmissao / mx if mx > 0 else 0
                    out.append(f"rgb({int(59+(249-59)*n)},{int(130+(115-130)*n)},{int(246+(22-246)*n)})")
                else: out.append("rgba(100,100,100,0.8)")
        return out

