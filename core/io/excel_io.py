"""
Leitura/escrita de arquivos Excel — migração e template.

Responsável por:
- Converter Excel → JSON DB (migração)
- Gerar template Excel de input
"""
from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════
#  ESTILOS COMUNS
# ═══════════════════════════════════════════════════════════════════
if OPENPYXL_AVAILABLE:
    _HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    _HDR_FILL = PatternFill(start_color="4c8061", end_color="4c8061", fill_type="solid")
    _OPT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    _REQ_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    _THIN_BORDER = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    _CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ═══════════════════════════════════════════════════════════════════
#  GERAÇÃO DE TEMPLATE EXCEL
# ═══════════════════════════════════════════════════════════════════

def gerar_template_excel(
    ano: Optional[int] = None,
    fatores_emissao: Optional[List[Dict]] = None,
) -> bytes:
    """Gera template Excel em memória para input de dados.

    Args:
        ano: Ano padrão a incluir.
        fatores_emissao: Lista de fatores para preencher dropdown.

    Returns:
        Conteúdo do arquivo .xlsx como bytes.

    Raises:
        ImportError: Se openpyxl não estiver instalado.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl é necessário para gerar templates Excel.")

    wb = Workbook()
    ano = ano or datetime.now().year

    # ─── ABA README ─────────────────────────────────────────────
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.sheet_properties.tabColor = "4c8061"

    readme_content = [
        ["TEMPLATE DE IMPORTAÇÃO — Calculadora de Emissões CMP"],
        [""],
        ["Este arquivo contém as abas necessárias para importar dados"],
        ["na Calculadora de Emissões de Carbono - CMP."],
        [""],
        ["INSTRUÇÕES:"],
        ["1. Preencha a aba 'Unidades' com as unidades produtivas da cadeia."],
        ["2. Preencha a aba 'Conexoes' com os fluxos entre unidades."],
        ["3. Preencha a aba 'Tecnologias' com tecnologias alternativas (opcional)."],
        ["4. A aba 'Fatores_Emissao' é referência; importe-a separadamente se necessário."],
        [""],
        ["CAMPOS OBRIGATÓRIOS estão marcados com (*) na linha de cabeçalho."],
        ["Campos com fundo laranja claro são obrigatórios."],
        ["Campos com fundo verde claro são opcionais."],
        [""],
        ["REGRAS:"],
        ["- ID_ELO deve ser único por unidade produtiva."],
        ["- Periodo deve ser um ano (ex: 2025)."],
        ["- Massas em toneladas (t)."],
        ["- Consumíveis e ConsumoEspecifico são listas separadas por ';'."],
        ["  Exemplo Consumíveis: DIESEL S10 (BRASIL);ELETRICIDADE"],
        ["  Exemplo ConsumoEspecifico: 0.5;1.2"],
        ["- Origem e Destino em Conexoes devem corresponder a ID_ELO existentes."],
        [""],
        [f"Ano padrão: {ano}"],
        [f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"],
        ["Versão do template: 1.0"],
    ]

    for i, row in enumerate(readme_content):
        cell = ws_readme.cell(row=i + 1, column=1, value=row[0] if row else "")
        if i == 0:
            cell.font = Font(name="Calibri", bold=True, size=14, color="4c8061")
        elif row and row[0].startswith("INSTRUÇÕES") or (row and row[0].startswith("REGRAS")):
            cell.font = Font(name="Calibri", bold=True, size=12, color="333333")

    ws_readme.column_dimensions["A"].width = 80

    # ─── ABA UNIDADES ───────────────────────────────────────────
    ws_u = wb.create_sheet("Unidades")
    ws_u.sheet_properties.tabColor = "0066CC"

    unidades_headers = [
        ("ID_ELO *", True),
        ("Nome *", True),
        ("Localizacao *", True),
        ("Periodo *", True),
        ("Input", False),
        ("MassaInput (t) *", True),
        ("Output", False),
        ("MassaOutput (t) *", True),
        ("Consumiveis", False),
        ("ConsumoEspecifico", False),
        ("TaxacaoFronteira", False),
        ("TaxacaoLocal", False),
        ("Tecnologia", False),
    ]
    _write_headers(ws_u, unidades_headers)

    # Exemplo de dados
    exemplo_unidade = [
        "U001", "Mina de Ferro", "Minas Gerais", str(ano),
        "Minério ROM", 1000.0, "Concentrado Fe", 500.0,
        "DIESEL S10 (BRASIL);ELETRICIDADE", "0.5;1.2",
        "FALSE", "FALSE", "",
    ]
    _write_example_row(ws_u, 3, exemplo_unidade)
    _auto_width(ws_u)

    # Validação para TaxacaoFronteira e TaxacaoLocal
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv_bool.error = "Valor deve ser TRUE ou FALSE"
    dv_bool.errorTitle = "Valor inválido"
    ws_u.add_data_validation(dv_bool)
    dv_bool.add(f"K2:L1000")

    # ─── ABA CONEXÕES ───────────────────────────────────────────
    ws_c = wb.create_sheet("Conexoes")
    ws_c.sheet_properties.tabColor = "CC6600"

    conexoes_headers = [
        ("Origem (ID_ELO) *", True),
        ("Destino (ID_ELO) *", True),
        ("Massa (t)", False),
        ("Label", False),
        ("Periodo *", True),
    ]
    _write_headers(ws_c, conexoes_headers)

    exemplo_conexao = ["U001", "U002", 500.0, "Fluxo Concentrado", "2025"]
    _write_example_row(ws_c, 3, exemplo_conexao)
    _auto_width(ws_c)

    # ─── ABA TECNOLOGIAS ────────────────────────────────────────
    ws_t = wb.create_sheet("Tecnologias")
    ws_t.sheet_properties.tabColor = "7C3AED"

    tec_headers = [
        ("ID *", True),
        ("Nome *", True),
        ("Insumos (nome1;nome2;...)", True),
        ("Fator_Consumo (fc1;fc2;...)", True),
    ]
    _write_headers(ws_t, tec_headers)

    exemplo_tec = ["TEC001", "Processo Convencional", "DIESEL S10 (BRASIL);ELETRICIDADE", "0.5;1.2"]
    _write_example_row(ws_t, 3, exemplo_tec)
    _auto_width(ws_t)

    # ─── ABA FATORES DE EMISSÃO (referência) ────────────────────
    ws_f = wb.create_sheet("Fatores_Emissao")
    ws_f.sheet_properties.tabColor = "059669"

    fatores_headers = [
        ("Grupo_Consumivel *", True),
        ("Consumivel *", True),
        ("Escopo *", True),
        ("Ano", False),
        ("Fator_Emissao *", True),
        ("kgCO2e_Unid *", True),
    ]
    _write_headers(ws_f, fatores_headers)

    # Preencher com fatores existentes se disponíveis
    if fatores_emissao:
        for i, f in enumerate(fatores_emissao[:200]):  # Limitar a 200 linhas
            row = i + 2
            ws_f.cell(row=row, column=1, value=f.get("grupo_consumivel", ""))
            ws_f.cell(row=row, column=2, value=f.get("consumivel", ""))
            ws_f.cell(row=row, column=3, value=f.get("escopo", ""))
            ws_f.cell(row=row, column=4, value=f.get("ano", ""))
            ws_f.cell(row=row, column=5, value=f.get("fator_emissao", 0.0))
            ws_f.cell(row=row, column=6, value=f.get("kgCO2e_unid", ""))
            for col in range(1, 7):
                ws_f.cell(row=row, column=col).border = _THIN_BORDER
    else:
        exemplo_fator = ["LAND TRANSPORTATION", "DIESEL S10 (BRASIL)", "SCOPE 1", str(ano), 2.35, "L"]
        _write_example_row(ws_f, 3, exemplo_fator)

    # Validação de escopo
    dv_escopo = DataValidation(
        type="list",
        formula1='"SCOPE 1,SCOPE 2,SCOPE 3"',
        allow_blank=True,
    )
    ws_f.add_data_validation(dv_escopo)
    dv_escopo.add(f"C2:C1000")

    _auto_width(ws_f)

    # ─── SALVAR ─────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════
#  EXPORTAÇÃO DE SESSÃO PARA EXCEL (dados preenchidos)
# ═══════════════════════════════════════════════════════════════════

def exportar_sessao_excel(
    unidades: List[Dict],
    conexoes: List[Dict],
    tecnologias: Optional[List[Dict]] = None,
    fatores_emissao: Optional[List[Dict]] = None,
    ano: Optional[int] = None,
) -> bytes:
    """Gera um arquivo Excel preenchido com os dados da sessão.

    Args:
        unidades: Lista de dicts (ou objetos com to_dict) das unidades.
        conexoes: Lista de dicts (ou objetos com to_dict) das conexões.
        tecnologias: Lista de dicts de tecnologias alternativas.
        fatores_emissao: Lista de dicts de fatores de emissão.
        ano: Ano de referência.

    Returns:
        Conteúdo do arquivo .xlsx como bytes.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl é necessário para exportar sessão em Excel.")

    wb = Workbook()
    ano = ano or datetime.now().year

    _DATA_FONT = Font(name="Calibri", size=10)

    # ─── ABA README ─────────────────────────────────────────────
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.sheet_properties.tabColor = "4c8061"
    readme_lines = [
        ["EXPORTAÇÃO DE SESSÃO — Calculadora de Emissões CMP"],
        [""],
        [f"Ano de referência: {ano}"],
        [f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"],
        [""],
        ["Abas: Unidades, Conexoes, Tecnologias, Fatores_Emissao"],
        ["Este arquivo pode ser reimportado na aplicação."],
    ]
    for i, row in enumerate(readme_lines):
        cell = ws_readme.cell(row=i + 1, column=1, value=row[0] if row else "")
        if i == 0:
            cell.font = Font(name="Calibri", bold=True, size=14, color="4c8061")
    ws_readme.column_dimensions["A"].width = 70

    # ─── ABA UNIDADES ───────────────────────────────────────────
    ws_u = wb.create_sheet("Unidades")
    ws_u.sheet_properties.tabColor = "0066CC"
    unidades_headers = [
        ("ID_ELO *", True),
        ("Nome *", True),
        ("Localizacao *", True),
        ("Periodo *", True),
        ("Input", False),
        ("MassaInput (t) *", True),
        ("Output", False),
        ("MassaOutput (t) *", True),
        ("Consumiveis", False),
        ("ConsumoEspecifico", False),
        ("TaxacaoFronteira", False),
        ("TaxacaoLocal", False),
        ("Tecnologia", False),
    ]
    _write_headers(ws_u, unidades_headers)

    for i, u in enumerate(unidades):
        d = u.to_dict() if hasattr(u, "to_dict") else (u if isinstance(u, dict) else vars(u))
        row = i + 3  # row 1=header, row 2=hints

        # Consumíveis: lista de dicts → "nome1;nome2"
        consumiveis_raw = d.get("Consumiveis") or d.get("consumiveis") or []
        if isinstance(consumiveis_raw, list):
            cons_nomes = ";".join(
                c["nome"] if isinstance(c, dict) else str(c) for c in consumiveis_raw
            )
        else:
            cons_nomes = str(consumiveis_raw)

        # ConsumoEspecifico: lista de floats → "0.5;1.2"
        ce_raw = d.get("ConsumoEspecifico") or d.get("consumo_especifico") or []
        if isinstance(ce_raw, list):
            cons_esp = ";".join(str(v) for v in ce_raw)
        else:
            cons_esp = str(ce_raw)

        # Tecnologia: pode ser obj, id ou None
        tec_val = d.get("Tecnologia") or d.get("tecnologia") or ""
        if isinstance(tec_val, dict):
            tec_val = tec_val.get("id", "")

        values = [
            d.get("ID_ELO") or d.get("id_elo", ""),
            d.get("Nome") or d.get("nome", ""),
            d.get("Localizacao") or d.get("localizacao", ""),
            d.get("Periodo") or d.get("periodo", str(ano)),
            d.get("Input") or d.get("input", ""),
            d.get("MassaInput") or d.get("massa_input", 0),
            d.get("Output") or d.get("output", ""),
            d.get("MassaOutput") or d.get("massa_output", 0),
            cons_nomes,
            cons_esp,
            str(d.get("TaxacaoFronteira", d.get("taxacao_fronteira", False))).upper(),
            str(d.get("TaxacaoLocal", d.get("taxacao_local", False))).upper(),
            tec_val,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_u.cell(row=row, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
    _auto_width(ws_u)

    # ─── ABA CONEXÕES ───────────────────────────────────────────
    ws_c = wb.create_sheet("Conexoes")
    ws_c.sheet_properties.tabColor = "CC6600"
    conexoes_headers = [
        ("Origem (ID_ELO) *", True),
        ("Destino (ID_ELO) *", True),
        ("Massa (t)", False),
        ("Label", False),
        ("Periodo *", True),
    ]
    _write_headers(ws_c, conexoes_headers)

    for i, c in enumerate(conexoes):
        d = c.to_dict() if hasattr(c, "to_dict") else (c if isinstance(c, dict) else vars(c))
        row = i + 3
        values = [
            d.get("origem", ""),
            d.get("destino", ""),
            d.get("massa", 0),
            d.get("label", "Fluxo"),
            d.get("periodo", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_c.cell(row=row, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
    _auto_width(ws_c)

    # ─── ABA TECNOLOGIAS ────────────────────────────────────────
    ws_t = wb.create_sheet("Tecnologias")
    ws_t.sheet_properties.tabColor = "7C3AED"
    tec_headers = [
        ("ID *", True),
        ("Nome *", True),
        ("Insumos (nome1;nome2;...)", True),
        ("Fator_Consumo (fc1;fc2;...)", True),
    ]
    _write_headers(ws_t, tec_headers)

    for i, t in enumerate(tecnologias or []):
        d = t.to_dict() if hasattr(t, "to_dict") else (t if isinstance(t, dict) else vars(t))
        row = i + 3
        insumos = d.get("insumos", [])
        nomes_ins = ";".join(ins.get("nome", "") if isinstance(ins, dict) else str(ins) for ins in insumos)
        fcs_ins = ";".join(str(ins.get("fator_consumo", 1.0)) if isinstance(ins, dict) else "1.0" for ins in insumos)
        values = [
            d.get("id", ""),
            d.get("nome", ""),
            nomes_ins,
            fcs_ins,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_t.cell(row=row, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
    _auto_width(ws_t)

    # ─── ABA FATORES DE EMISSÃO ─────────────────────────────────
    ws_f = wb.create_sheet("Fatores_Emissao")
    ws_f.sheet_properties.tabColor = "059669"
    fatores_headers = [
        ("Grupo_Consumivel *", True),
        ("Consumivel *", True),
        ("Escopo *", True),
        ("Ano", False),
        ("Fator_Emissao *", True),
        ("kgCO2e_Unid *", True),
    ]
    _write_headers(ws_f, fatores_headers)

    for i, f in enumerate(fatores_emissao or []):
        row = i + 3
        values = [
            f.get("grupo_consumivel", ""),
            f.get("consumivel", ""),
            f.get("escopo", ""),
            f.get("ano", ""),
            f.get("fator_emissao", 0.0),
            f.get("kgCO2e_unid", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_f.cell(row=row, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
    _auto_width(ws_f)

    # ─── SALVAR ─────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════
#  MIGRAÇÃO EXCEL → JSON
# ═══════════════════════════════════════════════════════════════════

def excel_to_json_db(filepath_or_bytes: Any) -> Dict[str, Any]:
    """Converte um arquivo Excel (template) para o formato JSON DB.

    Args:
        filepath_or_bytes: Caminho do arquivo ou bytes/BytesIO.

    Returns:
        Dict no formato do schema do banco de dados.
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("pandas é necessário para migração Excel → JSON.")

    from core.validation.schema import SCHEMA_VERSION

    sheets = pd.read_excel(filepath_or_bytes, sheet_name=None)

    result: Dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat(),
        "source": "excel_import",
        "anos_disponiveis": [],
        "fatores_emissao": [],
        "unidades": [],
        "conexoes": [],
        "tecnologias": [],
    }

    # Processar aba Fatores_Emissao
    if "Fatores_Emissao" in sheets:
        df = sheets["Fatores_Emissao"].dropna(how="all")
        # Normalizar nomes das colunas (remover * e espaços)
        df.columns = [c.replace(" *", "").replace("*", "").strip() for c in df.columns]
        col_map = {
            "Grupo_Consumivel": "grupo_consumivel",
            "Consumivel": "consumivel",
            "Escopo": "escopo",
            "Ano": "ano",
            "Fator_Emissao": "fator_emissao",
            "kgCO2e_Unid": "kgCO2e_unid",
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
        # Remover linhas com cabeçalho duplicado ou exemplo na row 2
        if "grupo_consumivel" in df.columns:
            df = df[df["grupo_consumivel"].notna()]
            df = df[~df["grupo_consumivel"].astype(str).str.strip().apply(_is_template_hint)]
        fatores_records = df.to_dict(orient="records")
        for rec in fatores_records:
            ano_val = rec.get("ano", None)
            if pd.isna(ano_val) or _is_template_hint(ano_val):
                rec.pop("ano", None)
            else:
                try:
                    rec["ano"] = int(float(ano_val))
                except (ValueError, TypeError):
                    rec.pop("ano", None)
        result["fatores_emissao"] = fatores_records

    # Processar aba Unidades
    anos_encontrados = set()
    if "Unidades" in sheets:
        df = sheets["Unidades"]
        # Normalizar nomes das colunas (remover * e espaços)
        df.columns = [c.replace(" *", "").replace("*", "").strip() for c in df.columns]
        # Remover linhas totalmente vazias e sem ID_ELO
        df = df.dropna(how="all")
        if "ID_ELO" in df.columns:
            df = df[df["ID_ELO"].notna()]
            df = df[~df["ID_ELO"].astype(str).str.strip().apply(_is_template_hint)]
        col_map = {
            "ID_ELO": "ID_ELO",
            "Nome": "Nome",
            "Localizacao": "Localizacao",
            "Periodo": "Periodo",
            "Input": "Input",
            "MassaInput (t)": "MassaInput",
            "MassaInput": "MassaInput",
            "Output": "Output",
            "MassaOutput (t)": "MassaOutput",
            "MassaOutput": "MassaOutput",
            "TaxacaoFronteira": "TaxacaoFronteira",
            "TaxacaoLocal": "TaxacaoLocal",
            "Tecnologia": "Tecnologia",
            "Consumiveis": "Consumiveis",
            "ConsumoEspecifico": "ConsumoEspecifico",
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

        for _, row in df.iterrows():
            periodo_ref = None
            try:
                periodo_ref = int(float(row.get("Periodo", datetime.now().year)))
            except (ValueError, TypeError):
                periodo_ref = None

            # Parse consumíveis (formato: "item1;item2")
            consumiveis_raw = str(row.get("Consumiveis", ""))
            consumo_raw = str(row.get("ConsumoEspecifico", ""))

            consumiveis = []
            consumo_especifico = []
            if consumiveis_raw and consumiveis_raw != "nan":
                nomes = [n.strip() for n in consumiveis_raw.split(";") if n.strip()]
                ces = [c.strip() for c in consumo_raw.split(";") if c.strip()] if consumo_raw != "nan" else []

                for i, nome in enumerate(nomes):
                    ce = _safe_float(ces[i], 0.0) if i < len(ces) else 0.0
                    # Buscar fator nos fatores importados
                    fator = 0.0
                    escopo = "1"
                    fator_resolvido = _resolver_fator_por_consumivel_ano(
                        nome,
                        result["fatores_emissao"],
                        periodo_ref,
                    )
                    if fator_resolvido:
                        fator = _safe_float(fator_resolvido.get("fator_emissao", 0.0), 0.0)
                        escopo = str(fator_resolvido.get("escopo", "1"))
                    consumiveis.append({"nome": nome, "fator": fator, "escopo": escopo})
                    consumo_especifico.append(ce)

            try:
                periodo = str(int(float(row.get("Periodo", datetime.now().year))))
            except (ValueError, TypeError):
                periodo = str(datetime.now().year)
            anos_encontrados.add(int(periodo))

            unidade = {
                "ID_ELO": str(row.get("ID_ELO", "")),
                "Nome": str(row.get("Nome", "")),
                "Localizacao": str(row.get("Localizacao", "")),
                "Periodo": periodo,
                "Input": str(row.get("Input", "")),
                "MassaInput": _safe_float(row.get("MassaInput", 0), 0.0),
                "Output": str(row.get("Output", "")),
                "MassaOutput": _safe_float(row.get("MassaOutput", 0), 0.0),
                "Consumiveis": consumiveis,
                "ConsumoEspecifico": consumo_especifico,
                "TaxacaoFronteira": _parse_bool(row.get("TaxacaoFronteira")),
                "TaxacaoLocal": _parse_bool(row.get("TaxacaoLocal")),
                "Tecnologia": str(row.get("Tecnologia", "")) if pd.notna(row.get("Tecnologia")) else None,
            }
            result["unidades"].append(unidade)

    # Processar aba Conexoes
    if "Conexoes" in sheets:
        df = sheets["Conexoes"].dropna(how="all")
        df.columns = [c.replace(" *", "").replace("*", "").replace("(ID_ELO) ", "").replace("(ID_ELO)", "").strip() for c in df.columns]
        col_map = {
            "Origem": "origem",
            "Destino": "destino",
            "Massa (t)": "massa",
            "Massa": "massa",
            "Label": "label",
            "Periodo": "periodo",
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
        if "origem" in df.columns:
            df = df[~df["origem"].astype(str).str.strip().apply(_is_template_hint)]
        if "destino" in df.columns:
            df = df[~df["destino"].astype(str).str.strip().apply(_is_template_hint)]
        for _, row in df.iterrows():
            if pd.notna(row.get("origem")) and pd.notna(row.get("destino")):
                result["conexoes"].append({
                    "origem": str(row["origem"]).strip(),
                    "destino": str(row["destino"]).strip(),
                    "massa": _safe_float(row.get("massa", 0), 0.0),
                    "label": str(row.get("label", "Fluxo")),
                    "periodo": str(row.get("periodo", "")) if pd.notna(row.get("periodo")) else "",
                })

    # Processar aba Tecnologias
    if "Tecnologias" in sheets:
        df = sheets["Tecnologias"].dropna(how="all")
        df.columns = [c.replace(" *", "").replace("*", "").strip() for c in df.columns]
        if "ID" in df.columns:
            df = df[~df["ID"].astype(str).str.strip().apply(_is_template_hint)]
        for _, row in df.iterrows():
            insumos_raw = str(row.get("Insumos (nome1;nome2;...)", row.get("Insumos", "")))
            fc_raw = str(row.get("Fator_Consumo (fc1;fc2;...)", row.get("Fator_Consumo", "")))

            insumos = []
            if insumos_raw and insumos_raw != "nan":
                nomes = [n.strip() for n in insumos_raw.split(";") if n.strip()]
                fcs = [c.strip() for c in fc_raw.split(";") if c.strip()] if fc_raw != "nan" else []
                for i, nome in enumerate(nomes):
                    fc = _safe_float(fcs[i], 1.0) if i < len(fcs) else 1.0
                    insumos.append({"nome": nome, "fator_consumo": fc})

            tec = {
                "id": str(row.get("ID", "")),
                "nome": str(row.get("Nome", "")),
                "insumos": insumos,
                "unidades": [],
            }
            result["tecnologias"].append(tec)

    result["anos_disponiveis"] = sorted(anos_encontrados) if anos_encontrados else [datetime.now().year]
    return result


# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════

def _parse_bool(val: Any) -> bool:
    """Parse boolean de vários formatos."""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().upper() in ("TRUE", "1", "SIM", "YES", "✅")
    return False


def _is_template_hint(val: Any) -> bool:
    """Retorna True para células de dica do template (obrigatório/opcional)."""
    if val is None:
        return False
    text = str(val).strip().lower()
    return text in {"(obrigatório)", "(opcional)"}


def _safe_float(val: Any, default: float = 0.0) -> float:
    """Converte para float sem lançar exceção em células textuais do template."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    if not text or _is_template_hint(text) or text.lower() == "nan":
        return default
    try:
        return float(text.replace(",", "."))
    except (ValueError, TypeError):
        return default


def _resolver_fator_por_consumivel_ano(
    consumivel: str,
    fatores: list[dict],
    ano: Optional[int],
) -> Optional[dict]:
    """Resolve fator por consumível priorizando ano exato e fallback global.

    Regras:
    1) Primeiro fator com consumível + ano exato
    2) Primeiro fator com consumível sem ano (global)
    3) Primeiro fator com consumível (qualquer ano)
    """
    nome = str(consumivel).strip().upper()
    if not nome:
        return None

    candidatos = [
        f for f in fatores
        if str(f.get("consumivel", "")).strip().upper() == nome
    ]
    if not candidatos:
        return None

    if ano is not None:
        for fator in candidatos:
            try:
                if int(fator.get("ano")) == int(ano):
                    return fator
            except (ValueError, TypeError):
                continue

    for fator in candidatos:
        ano_f = fator.get("ano", None)
        if ano_f is None or str(ano_f).strip() == "":
            return fator

    return candidatos[0]


def _write_headers(ws: Any, headers: list) -> None:
    """Escreve cabeçalhos com formatação."""
    for col_idx, (label, required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

        # Linha de dica (row 2)
        hint_cell = ws.cell(row=2, column=col_idx, value="(obrigatório)" if required else "(opcional)")
        hint_cell.font = Font(name="Calibri", italic=True, size=9, color="888888")
        hint_cell.fill = _REQ_FILL if required else _OPT_FILL
        hint_cell.alignment = _CENTER
        hint_cell.border = _THIN_BORDER


def _write_example_row(ws: Any, row: int, values: list) -> None:
    """Escreve linha de exemplo com formatação."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Calibri", italic=True, color="999999", size=10)
        cell.border = _THIN_BORDER


def _auto_width(ws: Any) -> None:
    """Ajusta largura das colunas automaticamente."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)
